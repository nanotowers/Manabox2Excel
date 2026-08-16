"""
manabox_to_excel.py  v2.1
--------------------------
Convierte una decklist O una colección de Manabox a:
  - Excel enriquecido con datos de Scryfall + gráfica de curva de maná
  - Reporte HTML interactivo con imágenes, estadísticas y filtros

Acepta DOS formatos de entrada:
  - .txt  → export de decklist de Manabox  ("1 Sol Ring (LTC) 123 *F*")
  - .csv  → export de colección/binder de Manabox (con cabeceras)

Estructura esperada:
    manabox_to_excel.py
    EDH\
        Sephiroth.txt       <- export de Manabox (mazo)
        Collection.csv      <- export de Manabox (colección / binder)
        Sephiroth.xlsx      <- generado automáticamente
        Sephiroth.html      <- reporte HTML generado automáticamente

Uso:
    python manabox_to_excel.py          # menú interactivo
    python manabox_to_excel.py --todos  # procesa todos los TXT y CSV

Novedades v2.1:
  - Parser CSV de Manabox con normalización de cabeceras y alias
  - Búsqueda por Scryfall ID / set+collector number (exacta, no fuzzy)
  - Descarga por lotes vía /cards/collection (75 cartas por request)
  - Columnas extra opcionales para CSV (condición, idioma, precio, binder)
"""

import re
import csv
import sys
import time
import json
import base64
import mimetypes
from datetime import datetime, timezone
from pathlib import Path
from collections import defaultdict

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint


# ── Configuración ─────────────────────────────────────────────────────────────

SCRYFALL_URL       = "https://api.scryfall.com/cards/named"
SCRYFALL_COLLECTION = "https://api.scryfall.com/cards/collection"
REQUEST_DELAY = 0.1
BATCH_SIZE    = 75          # límite duro de la API de Scryfall por request
MAX_REINTENTOS = 3          # reintentos ante 429 / 5xx / errores de red

# Scryfall EXIGE que las aplicaciones se identifiquen con User-Agent y Accept.
# Sin estas cabeceras la API responde 403 y todas las cartas salen "NO ENCONTRADA".
# Esta es la causa #1 de que un script que antes funcionaba deje de hacerlo.
USER_AGENT = "Manabox2Excel/2.2 (uso personal; https://scryfall.com/docs/api)"
HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept": "application/json",
}

# Sesión reutilizable: mantiene la conexión abierta y aplica las cabeceras
SESSION = requests.Session()
SESSION.headers.update(HEADERS)

# Control de mensajes de error: mostramos el primero con todo detalle y luego
# resumimos, para no llenar la consola con 9.000 líneas idénticas.
_ERRORES_VISTOS = set()


def reportar_error(contexto: str, detalle: str):
    """Imprime un error de red/API. Solo detalla la primera vez que ocurre."""
    clave = detalle[:80]
    if clave not in _ERRORES_VISTOS:
        _ERRORES_VISTOS.add(clave)
        print(f"\n  ✖ ERROR DE RED/API en {contexto}:")
        print(f"    {detalle}")
        print(f"    (los siguientes errores iguales se omiten)\n")
BASIC_LANDS   = {"swamp", "island", "plains", "mountain", "forest", "wastes"}

# Si True, las tierras básicas también se enriquecen con Scryfall.
# En v2.0 se saltaban (quedaban sin type_line y caían en la hoja "Other").
# Con descarga por lotes el costo es despreciable, así que ahora sí se cargan
# y aterrizan correctamente en la hoja "Land".
FETCH_BASIC_LANDS = True

# Columnas extra que SOLO se agregan cuando la fuente es un CSV de colección.
# Se añaden AL FINAL para no romper el orden documentado de las 20 columnas base.
# Ponlo en False si quieres que el CSV genere exactamente el mismo layout que el TXT.
INCLUDE_CSV_EXTRAS = True

# ── Modo de reporte HTML ──────────────────────────────────────────────────────
# "mazo"      → reporte de análisis: curva de maná, tipos, rareza, top cartas
# "coleccion" → navegador de inventario para compartir: filtros potentes,
#               identidad de comandante, vista tabla y lista de intercambio
#
# Se elige automáticamente: un CSV con muchas cartas es una colección; un TXT
# es un mazo. Se puede forzar con --coleccion o --mazo en la línea de comandos.
COLECCION_MIN_CARTAS = 200
MODO_FORZADO = None          # lo fija main() desde los argumentos
REPO_WEB = None              # ruta del repo web, la fija --web

# ── Personalización del HTML de colección ─────────────────────────────────────
# LOGO y BANNER aceptan DOS formatos:
#   - Ruta a un archivo local  → se incrusta en base64 dentro del HTML.
#     El archivo queda autocontenido: funciona sin internet y nunca se rompe.
#     Recomendado: logo hasta ~200 KB, banner hasta ~500 KB.
#   - URL que empiece por http → se enlaza tal cual (no engorda el archivo,
#     pero depende de que el hosting siga vivo).
#
# Si los dejas vacíos, el script busca automáticamente logo.* y banner.*
# junto al script o junto al CSV que estés procesando.
LOGO   = ""     # ej: "logo.png"  |  "https://i.imgur.com/xxxx.png"
BANNER = "https://i.imgur.com/bFaorlL.png"

TITULO_COLECCION = ""    # vacío = usa el nombre del archivo
SUBTITULO = "Colección de Magic: The Gathering · disponible para intercambio"

# Cómo se muestra el banner:
#   "hero"  → la imagen se ve completa, a lo ancho. Úsalo cuando el banner YA
#             trae el título dibujado dentro (así no se duplica el texto).
#             En este modo el <h1> se oculta automáticamente.
#   "fondo" → la imagen va de fondo, recortada, con un velo oscuro encima y el
#             título en texto sobre ella. Úsalo con imágenes sin texto.
BANNER_MODO = "hero"

EXTENSIONES_IMG = (".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg")

# Si es True y LOGO/BANNER son una URL, el script DESCARGA la imagen al generar
# y la incrusta en base64. Ventaja: el HTML queda autocontenido aunque la fuente
# sea externa, así que no se rompe si Imgur bloquea el hotlinking desde file://
# ni si el hosting desaparece. Si la descarga falla, se enlaza la URL tal cual.
DESCARGAR_URLS = True


def resolver_imagen(valor: str, carpetas: list[Path]) -> str:
    """
    Convierte LOGO/BANNER en algo usable dentro del HTML.
      - URL http(s)  → se descarga y se incrusta (o se enlaza si falla)
      - archivo local → data URI en base64 (queda incrustado en el HTML)
      - no encontrado → cadena vacía
    """
    valor = (valor or "").strip()

    if valor.startswith("data:"):
        return valor

    if valor.startswith(("http://", "https://")):
        if not DESCARGAR_URLS:
            return valor
        try:
            r = SESSION.get(valor, timeout=30)
            if r.status_code == 200 and r.content:
                mime = r.headers.get("Content-Type", "").split(";")[0].strip()
                if not mime.startswith("image/"):
                    mime = mimetypes.guess_type(valor)[0] or "image/png"
                b64 = base64.b64encode(r.content).decode("ascii")
                print(f"    imagen descargada e incrustada: {valor} "
                      f"(~{len(b64)//1024} KB en base64)")
                return f"data:{mime};base64,{b64}"
            print(f"    ⚠ HTTP {r.status_code} al descargar la imagen — se enlaza la URL")
        except Exception as e:
            print(f"    ⚠ No se pudo descargar la imagen ({type(e).__name__}) — se enlaza la URL")
        return valor

    candidatos = []
    if valor:
        p = Path(valor)
        candidatos.append(p)
        candidatos.extend(c / valor for c in carpetas)
    else:
        return ""

    for ruta in candidatos:
        try:
            if ruta.is_file():
                mime = mimetypes.guess_type(ruta.name)[0] or "image/png"
                b64 = base64.b64encode(ruta.read_bytes()).decode("ascii")
                kb = len(b64) // 1024
                print(f"    imagen incrustada: {ruta.name} (~{kb} KB en base64)")
                return f"data:{mime};base64,{b64}"
        except Exception as e:
            print(f"    ⚠ No se pudo leer {ruta}: {e}")
    print(f"    ⚠ No se encontró la imagen: {valor}")
    return ""


def buscar_imagen_auto(nombre_base: str, carpetas: list[Path]) -> str:
    """Busca logo.png / banner.jpg / etc. junto al script o al archivo de entrada."""
    for carpeta in carpetas:
        for ext in EXTENSIONES_IMG:
            ruta = carpeta / f"{nombre_base}{ext}"
            if ruta.is_file():
                return resolver_imagen(str(ruta), carpetas)
    return ""
# Ponlo en False si quieres que el CSV genere exactamente el mismo layout que el TXT.
INCLUDE_CSV_EXTRAS = True

COLUMNS = [
    "is_commander", "qty", "name", "foil",
    "mana_cost", "cmc", "type_line", "oracle_text",
    "power", "toughness", "loyalty",
    "keywords", "colors", "color_identity",
    "rarity", "set_code", "set_name",
    "flavor_text", "image_uri", "scryfall_uri",
]

CSV_EXTRA_COLUMNS = [
    "collector_number", "condition", "language",
    "purchase_price", "binder_name",
]

COLUMN_WIDTHS = {
    "is_commander": 12, "qty": 6,  "name": 28,  "foil": 6,
    "mana_cost": 14,    "cmc": 6,  "type_line": 30, "oracle_text": 55,
    "power": 7,   "toughness": 9,  "loyalty": 8,
    "keywords": 30, "colors": 10,  "color_identity": 14,
    "rarity": 10,  "set_code": 8,  "set_name": 20,
    "flavor_text": 40, "image_uri": 40, "scryfall_uri": 40,
    "collector_number": 10, "condition": 12, "language": 9,
    "purchase_price": 14, "binder_name": 18,
}

TYPE_COLORS = {
    "Commander":   "1A1A2E", "Creature":   "16213E",
    "Instant":     "0F3460", "Sorcery":    "533483",
    "Enchantment": "2D6A4F", "Artifact":   "4A4E69",
    "Land":        "6B4226", "Planeswalker": "B5451B",
    "Other":       "3D3D3D",
}

TYPE_ORDER = ["Commander", "Creature", "Instant", "Sorcery",
              "Enchantment", "Artifact", "Land", "Planeswalker", "Other"]

RARITY_BG = {
    "common": "FFFFFF", "uncommon": "C0C0C0",
    "rare": "FFD700",   "mythic": "FF8C00",
    "special": "9B59B6", "bonus": "3498DB",
}

# Colores MTG para símbolos
MANA_COLORS = {"W": "#F9FAF4", "U": "#0E68AB", "B": "#150B00",
               "R": "#D3202A", "G": "#00733E", "C": "#CCCCCC"}


# ── Parser Manabox ────────────────────────────────────────────────────────────

def parse_manabox(path: Path) -> list[dict]:
    cards, is_commander = [], False
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith("//"):
                is_commander = "commander" in line.lower()
                continue
            foil  = "*F*" in line
            match = re.match(r'^(\d+)\s+(.+?)(?:\s+\([^)]+\).*)?$', line)
            if not match:
                continue
            qty  = int(match.group(1))
            name = re.sub(r'\s+\([^)]+\).*', '', match.group(2)).strip()
            cards.append({"qty": qty, "name": name, "foil": foil, "is_commander": is_commander})
            is_commander = False
    return cards


# ── Parser CSV de Manabox ─────────────────────────────────────────────────────
#
# Manabox exporta la colección / los binders como CSV con cabecera. Ejemplo real:
#
#   Binder Name,Binder Type,Name,Set code,Set name,Collector number,Foil,Rarity,
#   Quantity,ManaBox ID,Scryfall ID,Purchase price,Misprint,Altered,Condition,
#   Language,Purchase price currency,Added
#
# Las cabeceras cambian según la versión de la app y según si exportas colección,
# binder o mazo, así que normalizamos y usamos alias en vez de posiciones fijas.

def _norm_header(h: str) -> str:
    """'Set code' -> 'setcode' ; 'Purchase price' -> 'purchaseprice'"""
    return re.sub(r'[^a-z0-9]', '', (h or "").strip().lower())


# alias_normalizado -> campo interno
CSV_FIELD_ALIASES = {
    "name": "name", "cardname": "name", "simplename": "name",
    "quantity": "qty", "qty": "qty", "count": "qty", "amount": "qty",
    "foil": "foil", "finish": "foil", "printing": "foil",
    "setcode": "set_code", "set": "set_code", "edition": "set_code",
    "setname": "set_name", "editionname": "set_name",
    "collectornumber": "collector_number", "cardnumber": "collector_number",
    "number": "collector_number",
    "scryfallid": "scryfall_id", "scryfalluuid": "scryfall_id",
    "rarity": "rarity",
    "condition": "condition",
    "language": "language", "lang": "language",
    "purchaseprice": "purchase_price", "price": "purchase_price",
    "purchasepricecurrency": "currency",
    "bindername": "binder_name", "binder": "binder_name", "folder": "binder_name",
    "bindertype": "binder_type",
    "iscommander": "is_commander", "commander": "is_commander",
}

FOIL_VALUES = {"foil", "etched", "etchedfoil", "glossy", "true", "yes", "1", "*f*"}


def parse_manabox_csv(path: Path) -> list[dict]:
    """
    Lee un CSV de Manabox y devuelve la misma estructura de dicts que
    parse_manabox(), más los campos extra disponibles en el CSV.

    Claves devueltas por carta:
        qty, name, foil, is_commander,
        scryfall_id, set_code, set_name, collector_number,
        condition, language, purchase_price, binder_name
    """
    cards = []
    # utf-8-sig quita el BOM que Manabox a veces mete al inicio del archivo
    with open(path, encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel          # fallback: coma
        reader = csv.reader(f, dialect)

        try:
            raw_headers = next(reader)
        except StopIteration:
            return cards

        # mapa: índice de columna -> campo interno
        idx_map = {}
        for i, h in enumerate(raw_headers):
            field = CSV_FIELD_ALIASES.get(_norm_header(h))
            if field and field not in idx_map.values():
                idx_map[i] = field

        if "name" not in idx_map.values():
            raise ValueError(
                f"El CSV '{path.name}' no tiene una columna de nombre reconocible.\n"
                f"Cabeceras encontradas: {raw_headers}"
            )

        for raw in reader:
            if not raw or not any(c.strip() for c in raw):
                continue
            rec = {}
            for i, field in idx_map.items():
                if i < len(raw):
                    rec[field] = (raw[i] or "").strip()

            name = rec.get("name", "").strip()
            if not name:
                continue

            # cantidad
            try:
                qty = int(float(rec.get("qty", "1") or 1))
            except ValueError:
                qty = 1
            if qty < 1:
                continue

            foil_raw = _norm_header(rec.get("foil", ""))
            is_cmd_raw = _norm_header(rec.get("is_commander", ""))

            cards.append({
                "qty": qty,
                "name": name,
                "foil": foil_raw in FOIL_VALUES,
                "is_commander": is_cmd_raw in {"true", "yes", "1", "y", "si", "x"},
                # extras del CSV
                "scryfall_id":      rec.get("scryfall_id", ""),
                "set_code":         rec.get("set_code", "").upper(),
                "set_name":         rec.get("set_name", ""),
                "collector_number": rec.get("collector_number", ""),
                "condition":        rec.get("condition", ""),
                "language":         rec.get("language", ""),
                "purchase_price":   rec.get("purchase_price", ""),
                "binder_name":      rec.get("binder_name", ""),
            })
    return cards


def parse_input(path: Path) -> list[dict]:
    """Dispatcher: elige el parser según la extensión del archivo."""
    if path.suffix.lower() == ".csv":
        return parse_manabox_csv(path)
    return parse_manabox(path)


# ── Scryfall ──────────────────────────────────────────────────────────────────

def _peticion(metodo: str, url: str, contexto: str, **kwargs):
    """
    Wrapper de requests con reintentos, backoff y errores visibles.
    Devuelve el objeto Response o None si agotó los reintentos.
    """
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            r = SESSION.request(metodo, url, timeout=30, **kwargs)

            if r.status_code == 200:
                return r

            # 429 = rate limit. Esperamos y reintentamos.
            if r.status_code == 429:
                espera = float(r.headers.get("Retry-After", 2 * intento))
                print(f"    (429 rate limit — esperando {espera}s)")
                time.sleep(espera)
                continue

            # 404 en /cards/named es normal: la carta simplemente no existe
            if r.status_code == 404:
                return None

            # 5xx = problema del servidor, vale la pena reintentar
            if r.status_code >= 500 and intento < MAX_REINTENTOS:
                time.sleep(2 * intento)
                continue

            cuerpo = ""
            try:
                cuerpo = r.json().get("details", "")
            except Exception:
                cuerpo = r.text[:200]
            reportar_error(contexto, f"HTTP {r.status_code} — {cuerpo}")
            return None

        except requests.exceptions.SSLError as e:
            reportar_error(contexto,
                f"SSL: {e}\n    → Suele ser antivirus/firewall/proxy interceptando HTTPS.")
            return None
        except requests.exceptions.ConnectionError as e:
            if intento < MAX_REINTENTOS:
                time.sleep(2 * intento)
                continue
            reportar_error(contexto,
                f"Conexión: {e}\n    → Sin salida a api.scryfall.com (DNS, firewall o proxy).")
            return None
        except requests.exceptions.Timeout:
            if intento < MAX_REINTENTOS:
                continue
            reportar_error(contexto, "Timeout tras 30s.")
            return None
        except Exception as e:
            reportar_error(contexto, f"{type(e).__name__}: {e}")
            return None
    return None


def fetch_scryfall(name: str) -> dict | None:
    """Búsqueda individual por nombre (fuzzy). Fallback del modo por lotes."""
    r = _peticion("GET", SCRYFALL_URL, f"búsqueda de '{name}'", params={"fuzzy": name})
    if r is None:
        return None
    try:
        return r.json()
    except Exception:
        return None


def diagnostico():
    """
    Comprueba la conectividad con Scryfall y muestra exactamente qué falla.
    Uso:  python manabox_to_excel.py --diagnostico
    """
    print("\n╔══════════════════════════════════════════╗")
    print("║      DIAGNÓSTICO DE CONEXIÓN SCRYFALL     ║")
    print("╚══════════════════════════════════════════╝\n")
    print(f"  Python       : {sys.version.split()[0]}")
    print(f"  requests     : {requests.__version__}")
    print(f"  User-Agent   : {USER_AGENT}\n")

    print("  [1/2] GET  /cards/named?fuzzy=Sol Ring")
    try:
        r = SESSION.get(SCRYFALL_URL, params={"fuzzy": "Sol Ring"}, timeout=30)
        print(f"        HTTP {r.status_code}")
        if r.status_code == 200:
            print(f"        OK → {r.json().get('name')} ({r.json().get('set','').upper()})")
        else:
            print(f"        Respuesta: {r.text[:300]}")
    except Exception as e:
        print(f"        FALLÓ → {type(e).__name__}: {e}")

    print("\n  [2/2] POST /cards/collection (2 cartas)")
    try:
        r = SESSION.post(
            SCRYFALL_COLLECTION,
            json={"identifiers": [{"name": "Sol Ring"}, {"name": "Lightning Bolt"}]},
            timeout=30,
        )
        print(f"        HTTP {r.status_code}")
        if r.status_code == 200:
            d = r.json()
            print(f"        OK → {len(d.get('data', []))} encontradas, "
                  f"{len(d.get('not_found', []))} no encontradas")
        else:
            print(f"        Respuesta: {r.text[:300]}")
    except Exception as e:
        print(f"        FALLÓ → {type(e).__name__}: {e}")

    print("\n  Si ambas dan HTTP 200, la API funciona y el problema está en el archivo.")
    print("  Si dan 403        → cabeceras rechazadas o IP bloqueada.")
    print("  Si dan SSLError   → antivirus/firewall interceptando HTTPS.")
    print("  Si dan Connection → sin salida a internet o proxy corporativo.\n")


def build_identifier(entry: dict) -> dict:
    """
    Construye el identificador más preciso disponible para /cards/collection.

    Prioridad:
      1. Scryfall ID       → carta exacta (edición, número, rareza correctos)
      2. set + collector_number → carta exacta
      3. name              → primera impresión que Scryfall considere canónica
    """
    sid = (entry.get("scryfall_id") or "").strip()
    if sid:
        return {"id": sid}
    set_code = (entry.get("set_code") or "").strip().lower()
    cn       = (entry.get("collector_number") or "").strip()
    if set_code and cn:
        return {"set": set_code, "collector_number": cn}
    return {"name": entry["name"]}


def fetch_scryfall_batch(identifiers: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Consulta /cards/collection en lotes de 75.
    Devuelve (cartas_encontradas, identificadores_no_encontrados).
    """
    found, not_found = [], []
    total_lotes = (len(identifiers) - 1) // BATCH_SIZE + 1
    for i in range(0, len(identifiers), BATCH_SIZE):
        chunk = identifiers[i:i + BATCH_SIZE]
        n_lote = i // BATCH_SIZE + 1

        r = _peticion("POST", SCRYFALL_COLLECTION,
                      f"lote {n_lote}", json={"identifiers": chunk})

        if r is None:
            not_found.extend(chunk)
            estado = "sin respuesta"
        else:
            try:
                payload = r.json()
                found.extend(payload.get("data", []))
                not_found.extend(payload.get("not_found", []))
                estado = f"{len(payload.get('data', []))} ok"
            except Exception as e:
                reportar_error(f"lote {n_lote}", f"JSON inválido: {e}")
                not_found.extend(chunk)
                estado = "JSON inválido"

        print(f"    lote {n_lote}/{total_lotes} "
              f"({min(i+BATCH_SIZE, len(identifiers))}/{len(identifiers)}) — {estado}")
        time.sleep(REQUEST_DELAY)
    return found, not_found


def extract_oracle(data: dict) -> str:
    if "oracle_text" in data:
        return data["oracle_text"]
    if "card_faces" in data:
        return "\n//\n".join(
            f"[{f.get('name','')}]\n{f.get('oracle_text','')}"
            for f in data["card_faces"]
        )
    return ""


def get_image(data: dict) -> str:
    if not data:
        return ""
    if "image_uris" in data:
        return data["image_uris"].get("normal", "")
    if "card_faces" in data:
        return data["card_faces"][0].get("image_uris", {}).get("normal", "")
    return ""


def enrich(entry: dict, data: dict | None, columns: list[str] | None = None) -> dict:
    columns = columns or COLUMNS
    row = {col: "" for col in columns}
    row.update({
        "is_commander": "YES" if entry["is_commander"] else "",
        "qty":  entry["qty"],
        "name": entry["name"],
        "foil": "✦" if entry["foil"] else "",
    })

    # Extras que vienen del propio CSV (no de Scryfall)
    for col in CSV_EXTRA_COLUMNS:
        if col in row and entry.get(col):
            row[col] = entry[col]
    if "purchase_price" in row and row["purchase_price"]:
        try:
            row["purchase_price"] = float(row["purchase_price"])
        except (TypeError, ValueError):
            pass

    if not data:
        # Sin datos de Scryfall conservamos lo que el CSV sí sabe
        for k in ("set_code", "set_name"):
            if not row.get(k) and entry.get(k):
                row[k] = entry[k]
        return row
    row["mana_cost"]     = data.get("mana_cost") or (data.get("card_faces",[{}])[0].get("mana_cost",""))
    row["cmc"]           = int(data["cmc"]) if data.get("cmc") is not None else ""
    row["type_line"]     = data.get("type_line", "")
    row["oracle_text"]   = extract_oracle(data)
    row["power"]         = data.get("power", "")
    row["toughness"]     = data.get("toughness", "")
    row["loyalty"]       = data.get("loyalty", "")
    row["keywords"]      = ", ".join(data.get("keywords", []))
    row["colors"]        = "".join(data.get("colors", []))
    row["color_identity"]= "".join(data.get("color_identity", []))
    row["rarity"]        = data.get("rarity", "")
    row["set_code"]      = data.get("set", "").upper()
    row["set_name"]      = data.get("set_name", "")
    row["flavor_text"]   = data.get("flavor_text", "")
    row["image_uri"]     = get_image(data)
    row["scryfall_uri"]  = data.get("scryfall_uri", "")
    if "collector_number" in row and not row["collector_number"]:
        row["collector_number"] = data.get("collector_number", "")
    return row


def sheet_for(row: dict) -> str:
    if row["is_commander"] == "YES":
        return "Commander"
    t = row.get("type_line", "")
    for key in ["Creature","Instant","Sorcery","Enchantment","Artifact","Land","Planeswalker"]:
        if key in t:
            return key
    return "Other"


# ── Excel helpers ─────────────────────────────────────────────────────────────

def hdr_font_color(hex_bg: str) -> str:
    r, g, b = int(hex_bg[0:2],16), int(hex_bg[2:4],16), int(hex_bg[4:6],16)
    return "000000" if (0.299*r + 0.587*g + 0.114*b)/255 > 0.5 else "FFFFFF"


def apply_header(ws, color_hex: str):
    bg = PatternFill("solid", start_color=color_hex)
    fc = hdr_font_color(color_hex)
    thin = Side(style="thin", color="CCCCCC")
    for cell in ws[1]:
        cell.font      = Font(name="Arial", bold=True, size=10, color=fc)
        cell.fill      = bg
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(bottom=thin)


def style_data_rows(ws, columns: list[str] | None = None):
    columns = columns or COLUMNS
    rarity_col = columns.index("rarity") + 1
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        rarity_val = ws.cell(row=ri, column=rarity_col).value or ""
        bg_hex  = RARITY_BG.get(rarity_val.lower(), "FFFFFF")
        alt_hex = "F9F9F9" if ri % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill      = PatternFill("solid", start_color=(bg_hex if cell.column == rarity_col and rarity_val else alt_hex))
    ws.freeze_panes  = "A2"
    ws.auto_filter.ref = ws.dimensions


def set_col_widths(ws, columns: list[str] | None = None):
    for idx, col in enumerate(columns or COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(col, 15)


# ── Hoja Resumen ──────────────────────────────────────────────────────────────

def write_summary(wb, groups: dict):
    ws = wb.create_sheet("Resumen", 0)
    ws.append(["Tipo", "Cartas", "% del Mazo", "CMC Promedio"])
    apply_header(ws, "1A1A2E")
    total = sum(r["qty"] for rows in groups.values() for r in rows if isinstance(r["qty"], int))
    for i, (sname, rows) in enumerate(groups.items(), start=2):
        count = sum(r["qty"] for r in rows if isinstance(r["qty"], int))
        cmcs  = [r["cmc"] for r in rows if isinstance(r.get("cmc"), int) and r["cmc"] != ""]
        avg   = round(sum(cmcs)/len(cmcs), 2) if cmcs else ""
        ws.append([sname, count, f"=B{i}/{total}" if total else 0, avg])
        ws.cell(row=i, column=3).number_format = "0.0%"
    last = len(groups) + 2
    ws.append(["TOTAL", f"=SUM(B2:B{last-1})", "", ""])
    ws.cell(row=last, column=1).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=last, column=2).font = Font(name="Arial", bold=True, size=10)
    for col, w in zip("ABCD", [18, 10, 14, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# ── Hoja Curva de Maná + Gráfica ─────────────────────────────────────────────

def write_mana_curve(wb, rows: list[dict]):
    ws = wb.create_sheet("Curva de Maná")

    # Excluir tierras y básicas del cálculo de curva
    non_lands = [
        r for r in rows
        if isinstance(r.get("cmc"), int)
        and "Land" not in r.get("type_line", "")
        and r["name"].lower() not in BASIC_LANDS
    ]

    # ── Tabla 1: Distribución por CMC ────────────────────────────────────────
    cmc_count  = defaultdict(int)
    cmc_by_type = defaultdict(lambda: defaultdict(int))
    types_seen  = set()

    for r in non_lands:
        cmc  = r["cmc"]
        stype = sheet_for(r)
        if stype in ("Land", "Commander", "Other"):
            stype = "Other"
        cmc_count[cmc]  += r["qty"]
        cmc_by_type[cmc][stype] += r["qty"]
        types_seen.add(stype)

    max_cmc  = max(cmc_count.keys(), default=0)
    cmc_range = list(range(0, min(max_cmc + 1, 10)))  # 0–9+

    types_ordered = [t for t in ["Creature","Instant","Sorcery","Enchantment","Artifact","Planeswalker","Other"] if t in types_seen]

    # Cabecera
    ws.cell(row=1, column=1).value = "CMC"
    ws.cell(row=1, column=1).font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws.cell(row=1, column=1).fill  = PatternFill("solid", start_color="1A1A2E")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    for ci, t in enumerate(types_ordered, start=2):
        cell = ws.cell(row=1, column=ci)
        cell.value = t
        cell.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color=TYPE_COLORS.get(t, "3D3D3D"))
        cell.alignment = Alignment(horizontal="center")

    total_col = len(types_ordered) + 2
    ws.cell(row=1, column=total_col).value = "Total"
    ws.cell(row=1, column=total_col).font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws.cell(row=1, column=total_col).fill  = PatternFill("solid", start_color="1A1A2E")
    ws.cell(row=1, column=total_col).alignment = Alignment(horizontal="center")

    # Filas de datos
    for ri, cmc in enumerate(cmc_range, start=2):
        label = f"{cmc}+" if cmc == 9 else str(cmc)
        # Para 9+ agrupamos todo lo que sea >= 9
        if cmc == 9:
            total_at_cmc = sum(v for k, v in cmc_count.items() if k >= 9)
            type_totals  = defaultdict(int)
            for k, tdict in cmc_by_type.items():
                if k >= 9:
                    for t, cnt in tdict.items():
                        type_totals[t] += cnt
        else:
            total_at_cmc = cmc_count.get(cmc, 0)
            type_totals  = cmc_by_type.get(cmc, {})

        ws.cell(row=ri, column=1).value = label
        ws.cell(row=ri, column=1).font  = Font(name="Arial", bold=True, size=9)
        ws.cell(row=ri, column=1).alignment = Alignment(horizontal="center")

        for ci, t in enumerate(types_ordered, start=2):
            val = type_totals.get(t, 0)
            cell = ws.cell(row=ri, column=ci)
            cell.value = val if val else ""
            cell.font  = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center")

        ws.cell(row=ri, column=total_col).value = total_at_cmc if total_at_cmc else ""
        ws.cell(row=ri, column=total_col).font  = Font(name="Arial", bold=True, size=9)
        ws.cell(row=ri, column=total_col).alignment = Alignment(horizontal="center")
        # Fila alternada
        bg = "F0F0F0" if ri % 2 == 0 else "FFFFFF"
        for ci2 in range(1, total_col + 1):
            ws.cell(row=ri, column=ci2).fill = PatternFill("solid", start_color=bg)

    # Fila total
    total_row = len(cmc_range) + 2
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws.cell(row=total_row, column=1).fill  = PatternFill("solid", start_color="1A1A2E")
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
    for ci, t in enumerate(types_ordered, start=2):
        total_t = sum(cmc_by_type[c].get(t, 0) for c in cmc_range[:-1]) + sum(v for k,v in cmc_by_type.items() if k >= 9 for tt,v in [(t, v.get(t,0))])
        val = sum(cmc_by_type[c].get(t, 0) for c in cmc_count if (c < 9 or c >= 9))
        # Simplificado: suma directa
        val2 = sum(r["qty"] for r in non_lands if sheet_for(r) == t or (t == "Other" and sheet_for(r) in ("Other","Commander")))
        ws.cell(row=total_row, column=ci).value = f"=SUM({get_column_letter(ci)}2:{get_column_letter(ci)}{total_row-1})"
        ws.cell(row=total_row, column=ci).font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        ws.cell(row=total_row, column=ci).fill  = PatternFill("solid", start_color="1A1A2E")
        ws.cell(row=total_row, column=ci).alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=total_col).value = f"=SUM({get_column_letter(total_col)}2:{get_column_letter(total_col)}{total_row-1})"
    ws.cell(row=total_row, column=total_col).font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws.cell(row=total_row, column=total_col).fill  = PatternFill("solid", start_color="1A1A2E")
    ws.cell(row=total_row, column=total_col).alignment = Alignment(horizontal="center")

    # Anchos de columna
    ws.column_dimensions["A"].width = 8
    for ci in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # ── Gráfica de barras apiladas ────────────────────────────────────────────
    chart = BarChart()
    chart.type          = "col"
    chart.grouping      = "stacked"
    chart.overlap       = 100
    chart.title         = "Curva de Maná"
    chart.y_axis.title  = "Cantidad de cartas"
    chart.x_axis.title  = "CMC"
    chart.style         = 10
    chart.width         = 22
    chart.height        = 14

    data_rows = len(cmc_range)  # filas de datos (sin total)

    for ci, t in enumerate(types_ordered, start=2):
        data   = Reference(ws, min_col=ci, max_col=ci, min_row=1, max_row=data_rows + 1)
        series = openpyxl.chart.Series(data, title_from_data=True)
        hex_c  = TYPE_COLORS.get(t, "888888")
        series.graphicalProperties.solidFill = hex_c
        series.graphicalProperties.line.solidFill = hex_c
        chart.series.append(series)

    cats = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)
    chart.set_categories(cats)
    chart.shape = 4

    # Anclar gráfica debajo de la tabla
    anchor_row = total_row + 2
    ws.add_chart(chart, f"A{anchor_row}")

    # ── Tabla 2: Estadísticas generales ──────────────────────────────────────
    stats_col = total_col + 2
    stats = [
        ("Estadísticas del Mazo", ""),
        ("Total cartas (sin tierras)", len(non_lands)),
        ("CMC promedio", round(sum(r["cmc"]*r["qty"] for r in non_lands if isinstance(r["cmc"],int)) /
                               max(sum(r["qty"] for r in non_lands), 1), 2)),
        ("CMC más común", max(cmc_count, key=cmc_count.get, default="-")),
        ("Cartas CMC ≤ 2", sum(r["qty"] for r in non_lands if isinstance(r["cmc"],int) and r["cmc"] <= 2)),
        ("Cartas CMC 3-4", sum(r["qty"] for r in non_lands if isinstance(r["cmc"],int) and 3 <= r["cmc"] <= 4)),
        ("Cartas CMC ≥ 5", sum(r["qty"] for r in non_lands if isinstance(r["cmc"],int) and r["cmc"] >= 5)),
        ("", ""),
        ("Tierras totales", sum(r["qty"] for r in rows if "Land" in r.get("type_line","") or r["name"].lower() in BASIC_LANDS)),
        ("Ramp (cartas)", sum(1 for r in rows if any(k in r.get("oracle_text","").lower() for k in ["add {","search your library for a basic land","mana","treasure"]))),
        ("Draw (cartas)", sum(1 for r in rows if "draw a card" in r.get("oracle_text","").lower())),
        ("Remoción (cartas)", sum(1 for r in rows if any(k in r.get("oracle_text","").lower() for k in ["destroy target","exile target","sacrifice a creature"]))),
    ]

    for si, (label, value) in enumerate(stats, start=1):
        lc = ws.cell(row=si, column=stats_col)
        vc = ws.cell(row=si, column=stats_col + 1)
        lc.value = label
        vc.value = value
        if si == 1:
            lc.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            lc.fill = PatternFill("solid", start_color="1A1A2E")
            vc.fill = PatternFill("solid", start_color="1A1A2E")
            ws.merge_cells(start_row=si, start_column=stats_col, end_row=si, end_column=stats_col+1)
            lc.alignment = Alignment(horizontal="center")
        else:
            lc.font = Font(name="Arial", size=9, bold=True)
            vc.font = Font(name="Arial", size=9)
            bg = "F0F0F0" if si % 2 == 0 else "FFFFFF"
            lc.fill = PatternFill("solid", start_color=bg)
            vc.fill = PatternFill("solid", start_color=bg)
            vc.alignment = Alignment(horizontal="center")

    ws.column_dimensions[get_column_letter(stats_col)].width   = 26
    ws.column_dimensions[get_column_letter(stats_col+1)].width = 10

    ws.freeze_panes = "A2"


# ── Generar HTML ──────────────────────────────────────────────────────────────

def generate_html(rows: list[dict], deck_name: str, output_path: Path):

    groups = {o: [] for o in TYPE_ORDER}
    for row in rows:
        groups[sheet_for(row)].append(row)
    groups = {k: v for k, v in groups.items() if v}

    # Estadísticas para charts
    non_lands = [r for r in rows if "Land" not in r.get("type_line","") and r["name"].lower() not in BASIC_LANDS and isinstance(r.get("cmc"), int)]
    cmc_dist  = defaultdict(int)
    for r in non_lands:
        cmc = min(r["cmc"], 9)
        cmc_dist[cmc] += r["qty"]
    cmc_labels = [str(i) if i < 9 else "9+" for i in range(10)]
    cmc_values = [cmc_dist.get(i, 0) for i in range(10)]

    type_labels = list(groups.keys())
    type_values = [sum(r["qty"] for r in v if isinstance(r["qty"],int)) for v in groups.values()]

    rarity_count = defaultdict(int)
    for r in rows:
        if r.get("rarity"):
            rarity_count[r["rarity"]] += r["qty"]

    total_cards = sum(r["qty"] for r in rows if isinstance(r["qty"],int))
    avg_cmc = round(sum(r["cmc"]*r["qty"] for r in non_lands)/max(sum(r["qty"] for r in non_lands),1), 2)
    lands   = sum(r["qty"] for r in rows if "Land" in r.get("type_line","") or r["name"].lower() in BASIC_LANDS)
    commander_name = next((r["name"] for r in rows if r.get("is_commander") == "YES"), deck_name)

    # ── Subtipos ──────────────────────────────────────────────────────────────
    # El subtipo vive después del guion largo en type_line:
    #   "Legendary Creature — Elf Druid"  → ["Elf", "Druid"]
    #   "Artifact — Equipment"            → ["Equipment"]
    #   DFC: se procesan ambas caras
    SUBTIPOS_COMPUESTOS = ["Time Lord"]      # los únicos de dos palabras

    def extraer_subtipos(type_line: str) -> list[str]:
        subs = []
        for cara in (type_line or "").split("//"):
            if "—" not in cara:
                continue
            cola = cara.split("—", 1)[1].strip()
            for compuesto in SUBTIPOS_COMPUESTOS:
                if compuesto in cola:
                    subs.append(compuesto)
                    cola = cola.replace(compuesto, " ")
            subs.extend(p.strip() for p in cola.split() if p.strip())
        # dedup conservando orden
        vistos, out = set(), []
        for s in subs:
            if s not in vistos:
                vistos.add(s)
                out.append(s)
        return out

    subtipo_count = defaultdict(int)
    for r in rows:
        for s in extraer_subtipos(r.get("type_line", "")):
            subtipo_count[s] += r["qty"] if isinstance(r["qty"], int) else 1
    subtipos_ordenados = sorted(subtipo_count.items(), key=lambda kv: kv[0].lower())

    # Construir cards JSON para JS
    cards_js = []
    for r in rows:
        cards_js.append({
            "name":       r["name"],
            "type":       sheet_for(r),
            "type_line":  r.get("type_line",""),
            "mana_cost":  r.get("mana_cost",""),
            "cmc":        r.get("cmc",""),
            "oracle":     r.get("oracle_text","").replace('"','\\"').replace("\n","\\n"),
            "rarity":     r.get("rarity",""),
            "image":      r.get("image_uri",""),
            "scryfall":   r.get("scryfall_uri",""),
            "foil":       r.get("foil",""),
            "qty":        r.get("qty",1),
            "keywords":   r.get("keywords",""),
            "power":      r.get("power",""),
            "toughness":  r.get("toughness",""),
            "set":        r.get("set_code",""),
            "is_commander": r.get("is_commander",""),
            "colors":     r.get("colors",""),          # colores impresos de la carta
            "ci":         r.get("color_identity",""),  # identidad de color (lo que importa en EDH)
            "subs":       extraer_subtipos(r.get("type_line","")),
        })

    cards_json = json.dumps(cards_js, ensure_ascii=False)

    type_colors_js = json.dumps({
        "Commander":"#1A1A2E","Creature":"#16213E","Instant":"#0F3460",
        "Sorcery":"#533483","Enchantment":"#2D6A4F","Artifact":"#4A4E69",
        "Land":"#6B4226","Planeswalker":"#B5451B","Other":"#3D3D3D"
    })

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{deck_name} — EDH Deck Report</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --bg: #0d0d0d; --surface: #1a1a2e; --surface2: #16213e;
    --accent: #e94560; --text: #e0e0e0; --text2: #aaaaaa;
    --border: #2a2a4a; --card-w: 180px;
    --common: #cccccc; --uncommon: #8ab4f8; --rare: #ffd700; --mythic: #ff8c00;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: var(--bg); color: var(--text); font-family: 'Segoe UI', Arial, sans-serif; }}

  /* NAV */
  nav {{ background: var(--surface); border-bottom: 1px solid var(--border);
        padding: 0 2rem; display: flex; align-items: center; gap: 2rem;
        position: sticky; top: 0; z-index: 100; height: 56px; }}
  nav .logo {{ font-size: 1.1rem; font-weight: 700; color: var(--accent); white-space: nowrap; }}
  nav .nav-links {{ display: flex; gap: 1.5rem; }}
  nav a {{ color: var(--text2); text-decoration: none; font-size: 0.875rem; transition: color .2s; }}
  nav a:hover {{ color: var(--text); }}

  /* HERO */
  .hero {{ background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
           padding: 3rem 2rem; text-align: center; border-bottom: 1px solid var(--border); }}
  .hero h1 {{ font-size: 2.5rem; font-weight: 800; color: #fff; margin-bottom: .5rem; }}
  .hero p  {{ color: var(--text2); font-size: 1rem; }}

  /* STATS BAR */
  .stats-bar {{ display: flex; gap: 1rem; padding: 1.5rem 2rem; flex-wrap: wrap;
               background: var(--surface); border-bottom: 1px solid var(--border); }}
  .stat {{ background: var(--surface2); border: 1px solid var(--border);
           border-radius: 12px; padding: .75rem 1.25rem; flex: 1; min-width: 120px; text-align: center; }}
  .stat-num  {{ font-size: 1.75rem; font-weight: 700; color: var(--accent); }}
  .stat-label{{ font-size: .75rem; color: var(--text2); margin-top: 2px; }}

  /* MAIN */
  main {{ max-width: 1400px; margin: 0 auto; padding: 2rem; }}
  section {{ margin-bottom: 3rem; }}
  h2 {{ font-size: 1.25rem; font-weight: 700; color: #fff; margin-bottom: 1.25rem;
       border-left: 4px solid var(--accent); padding-left: .75rem; }}

  /* CHARTS */
  .charts-grid {{ display: grid; grid-template-columns: 2fr 1fr 1fr; gap: 1.5rem; }}
  .chart-box {{ background: var(--surface); border: 1px solid var(--border);
               border-radius: 12px; padding: 1.25rem; }}
  .chart-box h3 {{ font-size: .9rem; color: var(--text2); margin-bottom: 1rem; }}

  /* FILTERS */
  .filters {{ display: flex; gap: .75rem; flex-wrap: wrap; margin-bottom: 1.5rem; align-items: center; }}
  .filter-btn {{ background: var(--surface); border: 1px solid var(--border); color: var(--text2);
                 padding: .4rem 1rem; border-radius: 20px; cursor: pointer;
                 font-size: .8rem; transition: all .2s; }}
  .filter-btn:hover, .filter-btn.active {{ border-color: var(--accent); color: #fff; background: var(--surface2); }}
  #search-input {{ background: var(--surface); border: 1px solid var(--border); color: var(--text);
                   padding: .4rem 1rem; border-radius: 20px; font-size: .85rem;
                   outline: none; min-width: 220px; }}
  #search-input:focus {{ border-color: var(--accent); }}

  /* FILTROS AVANZADOS: color + subtipo */
  .filter-panel {{ background: var(--surface); border: 1px solid var(--border);
                   border-radius: 12px; padding: 1rem; margin-bottom: 1.25rem;
                   display: flex; flex-wrap: wrap; gap: 1.25rem; align-items: center; }}
  .filter-group {{ display: flex; align-items: center; gap: .5rem; flex-wrap: wrap; }}
  .filter-label {{ font-size: .7rem; text-transform: uppercase; letter-spacing: .08em;
                   color: var(--text2); margin-right: .1rem; }}
  .pip {{ width: 34px; height: 34px; border-radius: 50%; cursor: pointer;
          border: 2px solid var(--border); display: flex; align-items: center;
          justify-content: center; font-weight: 800; font-size: .85rem;
          transition: all .15s; user-select: none; }}
  .pip:hover {{ transform: translateY(-2px); }}
  .pip.on {{ border-color: #fff; box-shadow: 0 0 0 2px var(--accent); }}
  .pip-W {{ background: #f8f6d8; color: #6b6242; }}
  .pip-U {{ background: #c1d7e9; color: #1f4d68; }}
  .pip-B {{ background: #4a4a4a; color: #d9d0d0; }}
  .pip-R {{ background: #e4a08a; color: #7a2f1c; }}
  .pip-G {{ background: #a3c095; color: #23502a; }}
  .pip-C {{ background: #cac5c0; color: #4a4a4a; }}
  .mini-select {{ background: var(--surface2); border: 1px solid var(--border); color: var(--text);
                  padding: .35rem .6rem; border-radius: 8px; font-size: .78rem; outline: none;
                  cursor: pointer; max-width: 260px; }}
  .mini-select:focus {{ border-color: var(--accent); }}
  .chip {{ background: var(--surface2); border: 1px solid var(--border); color: var(--text2);
           padding: .3rem .75rem; border-radius: 14px; cursor: pointer; font-size: .75rem;
           transition: all .15s; }}
  .chip:hover {{ color: var(--text); }}
  .chip.on {{ border-color: var(--accent); color: #fff; background: var(--accent); }}
  .btn-clear {{ background: transparent; border: 1px solid var(--accent); color: var(--accent);
                padding: .35rem .9rem; border-radius: 14px; cursor: pointer; font-size: .75rem; }}
  .btn-clear:hover {{ background: var(--accent); color: #fff; }}
  .result-count {{ font-size: .8rem; color: var(--text2); margin-bottom: 1rem; }}
  .result-count b {{ color: var(--accent); }}
  .load-more {{ display: block; margin: 2rem auto 0; background: var(--surface2);
                border: 1px solid var(--accent); color: var(--text); padding: .7rem 2rem;
                border-radius: 20px; cursor: pointer; font-size: .85rem; }}
  .load-more:hover {{ background: var(--accent); color: #fff; }}
  .no-results {{ text-align: center; color: var(--text2); padding: 3rem 1rem; }}
  .sort-select {{ background: var(--surface); border: 1px solid var(--border); color: var(--text);
                  padding: .4rem .75rem; border-radius: 20px; font-size: .8rem; outline: none; cursor: pointer; }}

  /* CARDS GRID */
  #cards-container {{ display: grid;
    grid-template-columns: repeat(auto-fill, minmax(var(--card-w), 1fr)); gap: 1rem; }}
  .mtg-card {{ background: var(--surface); border: 1px solid var(--border);
               border-radius: 10px; overflow: hidden; cursor: pointer;
               transition: transform .2s, border-color .2s; position: relative; }}
  .mtg-card:hover {{ transform: translateY(-4px); border-color: var(--accent); }}
  .mtg-card img {{ width: 100%; display: block; aspect-ratio: 63/88; object-fit: cover; background: #111; }}
  .mtg-card .no-img {{ width: 100%; aspect-ratio: 63/88; background: var(--surface2);
                       display: flex; flex-direction: column; align-items: center;
                       justify-content: center; gap: .5rem; padding: 1rem; text-align: center; }}
  .mtg-card .no-img .card-name-nimg {{ font-size: .75rem; font-weight: 600; color: var(--text); }}
  .mtg-card .no-img .card-type-nimg {{ font-size: .65rem; color: var(--text2); }}
  .card-footer {{ padding: .5rem .6rem; }}
  .card-footer .cname {{ font-size: .75rem; font-weight: 600; color: var(--text);
                         white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
  .card-footer .cmeta {{ font-size: .65rem; color: var(--text2); margin-top: 2px; }}
  .rarity-dot {{ position: absolute; top: 6px; right: 6px; width: 10px; height: 10px;
                 border-radius: 50%; border: 1px solid rgba(0,0,0,.4); }}
  .foil-badge {{ position: absolute; top: 6px; left: 6px; background: rgba(255,215,0,.85);
                 color: #000; font-size: .55rem; font-weight: 700; padding: 1px 5px;
                 border-radius: 4px; }}
  .commander-badge {{ position: absolute; bottom: 36px; left: 0; right: 0; text-align: center;
                      background: rgba(233,69,96,.9); color: #fff; font-size: .6rem;
                      font-weight: 700; padding: 2px 0; }}

  /* MODAL */
  .modal-overlay {{ display: none; position: fixed; inset: 0;
                    background: rgba(0,0,0,.85); z-index: 999;
                    align-items: center; justify-content: center; padding: 1rem; }}
  .modal-overlay.open {{ display: flex; }}
  .modal {{ background: var(--surface); border: 1px solid var(--border);
            border-radius: 16px; max-width: 700px; width: 100%;
            display: flex; gap: 1.5rem; padding: 1.5rem;
            max-height: 90vh; overflow-y: auto; }}
  .modal img {{ width: 220px; min-width: 220px; border-radius: 10px; object-fit: cover; }}
  .modal-info {{ flex: 1; }}
  .modal-info h3 {{ font-size: 1.2rem; font-weight: 700; margin-bottom: .5rem; }}
  .modal-info .mtype {{ color: var(--text2); font-size: .85rem; margin-bottom: .75rem; }}
  .modal-info .moracle {{ font-size: .85rem; line-height: 1.7; white-space: pre-wrap;
                          border-left: 3px solid var(--accent); padding-left: .75rem; margin-bottom: .75rem; }}
  .modal-info .mflavor {{ font-size: .8rem; color: var(--text2); font-style: italic; margin-bottom: .75rem; }}
  .tag {{ display: inline-block; padding: 2px 8px; border-radius: 4px;
          font-size: .7rem; font-weight: 600; margin: 2px; }}
  .modal-close {{ position: absolute; top: 1rem; right: 1rem; background: none;
                  border: none; color: var(--text2); font-size: 1.5rem; cursor: pointer; }}
  .modal-actions {{ display: flex; gap: .5rem; margin-top: .75rem; flex-wrap: wrap; }}
  .btn-scryfall {{ background: var(--accent); color: #fff; border: none; padding: .4rem 1rem;
                   border-radius: 6px; font-size: .8rem; cursor: pointer; text-decoration: none;
                   display: inline-block; }}

  /* SECTION TYPE HEADERS */
  .type-section {{ margin-bottom: 2rem; }}
  .type-header {{ display: flex; align-items: center; gap: .75rem; margin-bottom: .75rem;
                  padding: .5rem .75rem; border-radius: 8px; }}
  .type-header h3 {{ font-size: 1rem; font-weight: 700; color: #fff; }}
  .type-count {{ font-size: .8rem; color: rgba(255,255,255,.7); }}

  @media (max-width: 768px) {{
    .charts-grid {{ grid-template-columns: 1fr; }}
    .modal {{ flex-direction: column; }}
    .modal img {{ width: 100%; min-width: unset; }}
    nav .nav-links {{ display: none; }}
  }}
</style>
</head>
<body>

<nav>
  <span class="logo">⚔ {deck_name}</span>
  <div class="nav-links">
    <a href="#estadisticas">Estadísticas</a>
    <a href="#curva">Curva de Maná</a>
    <a href="#cartas">Cartas</a>
  </div>
</nav>

<div class="hero">
  <h1>{deck_name}</h1>
  <p>Commander: <strong>{commander_name}</strong> &nbsp;·&nbsp; EDH Deck Report</p>
</div>

<div class="stats-bar">
  <div class="stat"><div class="stat-num">{total_cards}</div><div class="stat-label">Total cartas</div></div>
  <div class="stat"><div class="stat-num">{lands}</div><div class="stat-label">Tierras</div></div>
  <div class="stat"><div class="stat-num">{total_cards - lands}</div><div class="stat-label">No-tierras</div></div>
  <div class="stat"><div class="stat-num">{avg_cmc}</div><div class="stat-label">CMC promedio</div></div>
  <div class="stat"><div class="stat-num">{rarity_count.get("mythic",0)}</div><div class="stat-label">Míticas</div></div>
  <div class="stat"><div class="stat-num">{rarity_count.get("rare",0)}</div><div class="stat-label">Raras</div></div>
</div>

<main>

  <!-- CHARTS -->
  <section id="curva">
    <h2>Curva de Maná &amp; Distribución</h2>
    <div class="charts-grid">
      <div class="chart-box">
        <h3>Curva de Maná (sin tierras)</h3>
        <canvas id="curvaChart" height="200"></canvas>
      </div>
      <div class="chart-box">
        <h3>Cartas por tipo</h3>
        <canvas id="tipoChart"></canvas>
      </div>
      <div class="chart-box">
        <h3>Por rareza</h3>
        <canvas id="rarezaChart"></canvas>
      </div>
    </div>
  </section>

  <!-- CARDS -->
  <section id="cartas">
    <h2>Cartas del Mazo</h2>
    <div class="filters" id="estadisticas">
      <input type="text" id="search-input" placeholder="Buscar carta...">
      <button class="filter-btn active" data-type="Todos">Todos</button>
      {"".join(f'<button class="filter-btn" data-type="{t}">{t}</button>' for t in type_labels)}
      <select class="sort-select" id="sort-select">
        <option value="type">Ordenar por tipo</option>
        <option value="cmc">Ordenar por CMC</option>
        <option value="name">Ordenar por nombre</option>
        <option value="rarity">Ordenar por rareza</option>
      </select>
    </div>

    <div class="filter-panel">
      <div class="filter-group">
        <span class="filter-label">Color</span>
        <div class="pip pip-W" data-color="W" title="Blanco">W</div>
        <div class="pip pip-U" data-color="U" title="Azul">U</div>
        <div class="pip pip-B" data-color="B" title="Negro">B</div>
        <div class="pip pip-R" data-color="R" title="Rojo">R</div>
        <div class="pip pip-G" data-color="G" title="Verde">G</div>
        <div class="pip pip-C" data-color="C" title="Incoloro">C</div>
        <select class="mini-select" id="color-mode" title="Cómo se combinan los colores marcados">
          <option value="any">Contiene alguno</option>
          <option value="all">Contiene todos</option>
          <option value="exact">Exactamente esos</option>
        </select>
        <select class="mini-select" id="color-source" title="Identidad de color = lo que importa en EDH">
          <option value="ci">Identidad de color</option>
          <option value="colors">Colores impresos</option>
        </select>
      </div>

      <div class="filter-group">
        <span class="filter-label">Cantidad</span>
        <button class="chip on" data-count="todos">Todas</button>
        <button class="chip" data-count="mono">Monocolor</button>
        <button class="chip" data-count="multi">Multicolor</button>
        <button class="chip" data-count="incoloro">Incoloras</button>
      </div>

      <div class="filter-group">
        <span class="filter-label">Subtipo</span>
        <select class="mini-select" id="subtype-select">
          <option value="">Todos los subtipos ({len(subtipos_ordenados)})</option>
          {"".join(f'<option value="{s}">{s} ({n})</option>' for s, n in subtipos_ordenados)}
        </select>
        <input type="text" class="mini-select" id="subtype-search" placeholder="o escribe: elf, snake..." style="min-width:170px">
      </div>

      <div class="filter-group">
        <span class="filter-label">Rareza</span>
        <select class="mini-select" id="rarity-select">
          <option value="">Todas</option>
          <option value="mythic">Mythic</option>
          <option value="rare">Rare</option>
          <option value="uncommon">Uncommon</option>
          <option value="common">Common</option>
        </select>
        <button class="chip" data-flag="foil">Solo foil</button>
        <button class="btn-clear" id="clear-filters">Limpiar filtros</button>
      </div>
    </div>

    <div class="result-count" id="result-count"></div>
    <div id="cards-container"></div>
  </section>

</main>

<!-- MODAL -->
<div class="modal-overlay" id="modal-overlay">
  <div class="modal" id="modal-content"></div>
</div>

<script>
const CARDS = {cards_json};
const TYPE_COLORS = {type_colors_js};
const RARITY_COLORS = {{common:"#aaaaaa",uncommon:"#8ab4f8",rare:"#ffd700",mythic:"#ff8c00",special:"#9b59b6"}};

// ── Charts ────────────────────────────────────────────────────────────────────
const chartDefaults = {{
  plugins: {{ legend: {{ labels: {{ color: "#aaaaaa", font: {{ size: 11 }} }} }} }},
  scales:  {{ x: {{ ticks: {{ color:"#aaaaaa" }}, grid: {{ color:"#2a2a4a" }} }},
              y: {{ ticks: {{ color:"#aaaaaa" }}, grid: {{ color:"#2a2a4a" }} }} }}
}};

new Chart(document.getElementById("curvaChart"), {{
  type: "bar",
  data: {{
    labels: {json.dumps(cmc_labels)},
    datasets: [{{ label: "Cartas", data: {json.dumps(cmc_values)},
      backgroundColor: CARDS.map ? "#0f3460" : "#0f3460",
      backgroundColor: Array({len(cmc_values)}).fill(0).map((_,i) => i<=2?"#2D6A4F":i<=4?"#0F3460":"#533483"),
      borderRadius: 4 }}]
  }},
  options: {{ ...chartDefaults, plugins: {{ legend: {{ display: false }} }} }}
}});

new Chart(document.getElementById("tipoChart"), {{
  type: "doughnut",
  data: {{
    labels: {json.dumps(type_labels)},
    datasets: [{{ data: {json.dumps(type_values)},
      backgroundColor: {json.dumps(type_labels)}.map(t => TYPE_COLORS[t] || "#3d3d3d"),
      borderWidth: 1, borderColor: "#0d0d0d" }}]
  }},
  options: {{ plugins: {{ legend: {{ position:"right", labels: {{ color:"#aaaaaa", font:{{size:10}}, boxWidth:12 }} }} }} }}
}});

const rarezaLabels = {json.dumps(list(rarity_count.keys()))};
const rarezaValues = {json.dumps(list(rarity_count.values()))};
new Chart(document.getElementById("rarezaChart"), {{
  type: "doughnut",
  data: {{
    labels: rarezaLabels,
    datasets: [{{ data: rarezaValues,
      backgroundColor: rarezaLabels.map(r => RARITY_COLORS[r] || "#aaaaaa"),
      borderWidth: 1, borderColor: "#0d0d0d" }}]
  }},
  options: {{ plugins: {{ legend: {{ position:"right", labels: {{ color:"#aaaaaa", font:{{size:10}}, boxWidth:12 }} }} }} }}
}});

// ── Cards rendering ───────────────────────────────────────────────────────────
const TYPE_ORDER_HTML = ["Commander","Creature","Instant","Sorcery","Enchantment","Artifact","Land","Planeswalker","Other"];
let activeType = "Todos";
let activeSort = "type";

// ── Estado de los filtros avanzados ───────────────────────────────────────────
const PAGE_SIZE = 300;          // cartas renderizadas por tanda
let shownCount   = PAGE_SIZE;
let activeColors = new Set();   // W U B R G C
let colorMode    = "any";       // any | all | exact
let colorSource  = "ci";        // ci (identidad, la que importa en EDH) | colors
let colorCount   = "todos";     // todos | mono | multi | incoloro
let onlyFoil     = false;

function colorMatch(c) {{
  const cs  = (colorSource === "ci" ? c.ci : c.colors) || "";
  const arr = cs.split("");
  const n   = arr.length;

  // Filtro por cantidad de colores
  if (colorCount === "mono"     && n !== 1) return false;
  if (colorCount === "multi"    && n < 2)   return false;
  if (colorCount === "incoloro" && n !== 0) return false;

  if (activeColors.size === 0) return true;

  const sel  = [...activeColors];
  const conC = sel.includes("C");
  const cols = sel.filter(x => x !== "C");

  if (colorMode === "any") {{
    if (conC && n === 0) return true;
    return cols.some(x => arr.includes(x));
  }}
  if (colorMode === "all") {{
    if (!cols.length) return conC ? n === 0 : true;
    return cols.every(x => arr.includes(x));
  }}
  // exact: la identidad debe ser exactamente la marcada, ni un color de más
  if (!cols.length) return n === 0;
  return n === cols.length && cols.every(x => arr.includes(x));
}}

function rarityDotColor(rarity) {{
  return RARITY_COLORS[rarity] || "#555";
}}

function renderCard(c) {{
  const imgHtml = c.image
    ? `<img src="${{c.image}}" alt="${{c.name}}" loading="lazy">`
    : `<div class="no-img">
         <div style="font-size:2rem">🃏</div>
         <div class="card-name-nimg">${{c.name}}</div>
         <div class="card-type-nimg">${{c.type_line || c.type}}</div>
       </div>`;
  return `
    <div class="mtg-card" onclick="openModal(${{JSON.stringify(c).replace(/'/g,"&apos;")}})">
      ${{c.foil ? '<div class="foil-badge">✦ FOIL</div>' : ""}}
      ${{c.rarity ? `<div class="rarity-dot" style="background:${{rarityDotColor(c.rarity)}}"></div>` : ""}}
      ${{c.is_commander === "YES" ? '<div class="commander-badge">COMMANDER</div>' : ""}}
      ${{imgHtml}}
      <div class="card-footer">
        <div class="cname">${{c.qty > 1 ? c.qty+"x " : ""}}${{c.name}}</div>
        <div class="cmeta">${{c.mana_cost || ""}} ${{c.cmc !== "" ? "· CMC "+c.cmc : ""}}</div>
      </div>
    </div>`;
}}

function filterAndRender(resetPage) {{
  if (resetPage !== false) shownCount = PAGE_SIZE;

  let filtered = activeType === "Todos" ? [...CARDS] : CARDS.filter(c => c.type === activeType);

  // Texto libre: nombre u oracle
  const q = document.getElementById("search-input").value.toLowerCase().trim();
  if (q) filtered = filtered.filter(c => c.name.toLowerCase().includes(q) || (c.oracle||"").toLowerCase().includes(q));

  // Color
  filtered = filtered.filter(colorMatch);

  // Subtipo: el select manda; si está en "Todos", usa el campo de texto
  let sub = document.getElementById("subtype-select").value;
  if (!sub) sub = document.getElementById("subtype-search").value.trim();
  if (sub) {{
    const needle = sub.toLowerCase();
    filtered = filtered.filter(c => (c.subs||[]).some(s => s.toLowerCase().includes(needle)));
  }}

  // Rareza
  const rar = document.getElementById("rarity-select").value;
  if (rar) filtered = filtered.filter(c => c.rarity === rar);

  // Solo foil
  if (onlyFoil) filtered = filtered.filter(c => c.foil);

  // Sort
  const sortFns = {{
    cmc: (a,b) => (a.cmc||99) - (b.cmc||99),
    name: (a,b) => a.name.localeCompare(b.name),
    rarity: (a,b) => {{
      const order = {{mythic:0,rare:1,uncommon:2,common:3}};
      return (order[a.rarity]||9) - (order[b.rarity]||9);
    }},
    type: (a,b) => (TYPE_ORDER_HTML.indexOf(a.type)||99) - (TYPE_ORDER_HTML.indexOf(b.type)||99)
  }};
  filtered.sort(sortFns[activeSort] || sortFns.type);

  // Contador
  const copias = filtered.reduce((s,c) => s + (c.qty||1), 0);
  document.getElementById("result-count").innerHTML =
    `<b>${{filtered.length}}</b> cartas distintas · <b>${{copias}}</b> copias` +
    (filtered.length > shownCount ? ` · mostrando las primeras ${{shownCount}}` : "");

  // Paginado: con colecciones grandes renderizar 14.000 imágenes cuelga el navegador
  const visibles = filtered.slice(0, shownCount);
  const container = document.getElementById("cards-container");

  if (!visibles.length) {{
    container.style.display = "block";
    container.innerHTML = `<div class="no-results">Ninguna carta coincide con estos filtros.</div>`;
    return;
  }}

  if (activeSort === "type" && activeType === "Todos") {{
    // Agrupar por tipo
    const byType = {{}};
    visibles.forEach(c => {{ if(!byType[c.type]) byType[c.type]=[]; byType[c.type].push(c); }});
    container.innerHTML = TYPE_ORDER_HTML.filter(t => byType[t]).map(t => {{
      const bg = TYPE_COLORS[t] || "#3d3d3d";
      return `<div class="type-section" style="grid-column:1/-1">
        <div class="type-header" style="background:${{bg}}20;border-left:4px solid ${{bg}}">
          <h3>${{t}}</h3>
          <span class="type-count">${{byType[t].reduce((s,c)=>s+c.qty,0)}} cartas</span>
        </div>
        <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(var(--card-w),1fr));gap:1rem">
          ${{byType[t].map(renderCard).join("")}}
        </div>
      </div>`;
    }}).join("");
    container.style.display = "block";
  }} else {{
    container.style.display = "grid";
    container.innerHTML = visibles.map(renderCard).join("");
  }}

  // Botón "cargar más"
  if (filtered.length > shownCount) {{
    const btn = document.createElement("button");
    btn.className = "load-more";
    btn.textContent = `Cargar ${{Math.min(PAGE_SIZE, filtered.length - shownCount)}} más ` +
                      `(quedan ${{filtered.length - shownCount}})`;
    btn.onclick = () => {{ shownCount += PAGE_SIZE; filterAndRender(false); }};
    container.appendChild(btn);
  }}
}}

// Filters
document.querySelectorAll(".filter-btn").forEach(btn => {{
  btn.addEventListener("click", () => {{
    document.querySelectorAll(".filter-btn").forEach(b => b.classList.remove("active"));
    btn.classList.add("active");
    activeType = btn.dataset.type;
    filterAndRender();
  }});
}});
document.getElementById("search-input").addEventListener("input", filterAndRender);
document.getElementById("sort-select").addEventListener("change", e => {{
  activeSort = e.target.value;
  filterAndRender();
}});

// ── Eventos de filtros avanzados ──────────────────────────────────────────────
document.querySelectorAll(".pip").forEach(pip => {{
  pip.addEventListener("click", () => {{
    const col = pip.dataset.color;
    if (activeColors.has(col)) {{ activeColors.delete(col); pip.classList.remove("on"); }}
    else                       {{ activeColors.add(col);    pip.classList.add("on"); }}
    filterAndRender();
  }});
}});

document.getElementById("color-mode").addEventListener("change", e => {{
  colorMode = e.target.value; filterAndRender();
}});
document.getElementById("color-source").addEventListener("change", e => {{
  colorSource = e.target.value; filterAndRender();
}});

document.querySelectorAll(".chip[data-count]").forEach(chip => {{
  chip.addEventListener("click", () => {{
    document.querySelectorAll(".chip[data-count]").forEach(c => c.classList.remove("on"));
    chip.classList.add("on");
    colorCount = chip.dataset.count;
    filterAndRender();
  }});
}});

const foilChip = document.querySelector('.chip[data-flag="foil"]');
foilChip.addEventListener("click", () => {{
  onlyFoil = !onlyFoil;
  foilChip.classList.toggle("on", onlyFoil);
  filterAndRender();
}});

// El select y el campo de texto de subtipo se limpian mutuamente
document.getElementById("subtype-select").addEventListener("change", () => {{
  if (document.getElementById("subtype-select").value) document.getElementById("subtype-search").value = "";
  filterAndRender();
}});
document.getElementById("subtype-search").addEventListener("input", () => {{
  if (document.getElementById("subtype-search").value) document.getElementById("subtype-select").value = "";
  filterAndRender();
}});
document.getElementById("rarity-select").addEventListener("change", filterAndRender);

document.getElementById("clear-filters").addEventListener("click", () => {{
  activeColors.clear();
  document.querySelectorAll(".pip").forEach(p => p.classList.remove("on"));
  colorMode = "any";   document.getElementById("color-mode").value = "any";
  colorSource = "ci";  document.getElementById("color-source").value = "ci";
  colorCount = "todos";
  document.querySelectorAll(".chip[data-count]").forEach(c => c.classList.remove("on"));
  document.querySelector('.chip[data-count="todos"]').classList.add("on");
  onlyFoil = false; foilChip.classList.remove("on");
  document.getElementById("subtype-select").value = "";
  document.getElementById("subtype-search").value = "";
  document.getElementById("rarity-select").value = "";
  document.getElementById("search-input").value = "";
  activeType = "Todos";
  document.querySelectorAll(".filter-btn").forEach(b => b.classList.remove("active"));
  document.querySelector('.filter-btn[data-type="Todos"]').classList.add("active");
  filterAndRender();
}});

// ── Modal ─────────────────────────────────────────────────────────────────────
function openModal(card) {{
  if (typeof card === "string") card = JSON.parse(card);
  const ptHtml = card.power ? `<span class="tag" style="background:#16213e;color:#fff">${{card.power}}/${{card.toughness}}</span>` : "";
  const rarTag = card.rarity ? `<span class="tag" style="background:${{rarityDotColor(card.rarity)}}20;color:${{rarityDotColor(card.rarity)}}">${{card.rarity}}</span>` : "";
  const foilTag = card.foil ? `<span class="tag" style="background:#ffd70020;color:#ffd700">✦ FOIL</span>` : "";
  const keyTags = card.keywords ? card.keywords.split(",").map(k => `<span class="tag" style="background:#0f346020;color:#8ab4f8">${{k.trim()}}</span>`).join("") : "";
  const imgHtml = card.image
    ? `<img src="${{card.image}}" alt="${{card.name}}">`
    : `<div style="width:220px;min-width:220px;background:#16213e;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:4rem">🃏</div>`;

  document.getElementById("modal-content").innerHTML = `
    ${{imgHtml}}
    <div class="modal-info">
      <h3>${{card.qty > 1 ? card.qty+"x " : ""}}${{card.name}} ${{card.is_commander === "YES" ? "👑" : ""}}</h3>
      <div class="mtype">${{card.type_line || card.type}} ${{card.mana_cost ? "· "+card.mana_cost : ""}} ${{card.cmc !== "" ? "· CMC "+card.cmc : ""}}</div>
      <div style="margin-bottom:.75rem">${{rarTag}}${{foilTag}}${{ptHtml}}${{keyTags}}</div>
      ${{card.oracle ? `<div class="moracle">${{card.oracle.replace(/\\n/g,"<br>")}}</div>` : ""}}
      ${{card.flavor ? `<div class="mflavor">${{card.flavor}}</div>` : ""}}
      <div style="font-size:.75rem;color:#666;margin-bottom:.5rem">Set: ${{card.set || "—"}}</div>
      <div class="modal-actions">
        ${{card.scryfall ? `<a href="${{card.scryfall}}" target="_blank" class="btn-scryfall">Ver en Scryfall ↗</a>` : ""}}
      </div>
    </div>`;
  document.getElementById("modal-overlay").classList.add("open");
}}

document.getElementById("modal-overlay").addEventListener("click", e => {{
  if (e.target === document.getElementById("modal-overlay"))
    document.getElementById("modal-overlay").classList.remove("open");
}});
document.addEventListener("keydown", e => {{
  if (e.key === "Escape") document.getElementById("modal-overlay").classList.remove("open");
}});

// Initial render
filterAndRender();
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)


# ── Procesar un TXT ───────────────────────────────────────────────────────────

def cache_key(entry: dict) -> str:
    """
    Clave de deduplicación para el cache de Scryfall.

    Se usa el Scryfall ID o set+número cuando existen (CSV): así una misma carta
    en dos ediciones distintas NO se colapsa en una sola entrada, que es
    exactamente lo que pasaba al indexar solo por nombre.
    """
    sid = (entry.get("scryfall_id") or "").strip()
    if sid:
        return f"id:{sid}"
    set_code = (entry.get("set_code") or "").strip().lower()
    cn       = (entry.get("collector_number") or "").strip()
    if set_code and cn:
        return f"cn:{set_code}:{cn}"
    return f"name:{entry['name'].lower()}"


def build_cache(uniques: dict) -> dict:
    """
    Descarga los datos de Scryfall de todas las cartas únicas.
    Devuelve {cache_key: data|None}.

    Estrategia:
      1. Un solo POST por cada 75 cartas a /cards/collection (rapidísimo)
      2. Las que no aparezcan se reintentan una a una con búsqueda fuzzy
    """
    cache = {}
    pending = {}

    for key, entry in uniques.items():
        if not FETCH_BASIC_LANDS and entry["name"].lower() in BASIC_LANDS:
            cache[key] = None
            continue
        pending[key] = entry

    if not pending:
        return cache

    keys        = list(pending.keys())
    identifiers = [build_identifier(pending[k]) for k in keys]

    print(f"  Consultando Scryfall por lotes de {BATCH_SIZE}...")
    found, not_found = fetch_scryfall_batch(identifiers)

    # Indexamos lo encontrado para poder reasociarlo con cada entrada
    by_id, by_setcn, by_name = {}, {}, {}
    for d in found:
        if d.get("id"):
            by_id[d["id"]] = d
        if d.get("set") and d.get("collector_number"):
            by_setcn[f"{d['set'].lower()}:{d['collector_number']}"] = d
        if d.get("name"):
            by_name.setdefault(d["name"].lower(), d)

    faltantes = []
    for key in keys:
        entry = pending[key]
        ident = build_identifier(entry)
        data  = None
        if "id" in ident:
            data = by_id.get(ident["id"])
        elif "set" in ident:
            data = by_setcn.get(f"{ident['set']}:{ident['collector_number']}")
        else:
            data = by_name.get(entry["name"].lower())
        if data is None:
            data = by_name.get(entry["name"].lower())
        if data is None:
            faltantes.append(key)
        cache[key] = data

    # Reintento individual con fuzzy para lo que no salió en el lote
    if faltantes:
        print(f"  Reintentando {len(faltantes)} carta(s) con búsqueda fuzzy...")
        for key in faltantes:
            entry = pending[key]
            data  = fetch_scryfall(entry["name"])
            cache[key] = data
            estado = "ok" if data else "NO ENCONTRADA"
            print(f"    - {entry['name']}: {estado}")
            time.sleep(REQUEST_DELAY)

    return cache




# ══════════════════════════════════════════════════════════════════════════════
# PLANTILLA HTML — MODO COLECCIÓN
# ══════════════════════════════════════════════════════════════════════════════


# Subtipos de dos palabras (prácticamente el único caso en Magic)
SUBTIPOS_COMPUESTOS = ["Time Lord"]


def extraer_subtipos(type_line: str) -> list[str]:
    """'Legendary Creature — Elf Druid' → ['Elf', 'Druid']. Procesa ambas caras."""
    subs = []
    for cara in (type_line or "").split("//"):
        if "—" not in cara:
            continue
        cola = cara.split("—", 1)[1].strip()
        for compuesto in SUBTIPOS_COMPUESTOS:
            if compuesto in cola:
                subs.append(compuesto)
                cola = cola.replace(compuesto, " ")
        subs.extend(p.strip() for p in cola.split() if p.strip())
    vistos, out = set(), []
    for s in subs:
        if s not in vistos:
            vistos.add(s)
            out.append(s)
    return out


def img_small(uri: str) -> str:
    """Scryfall sirve la misma imagen en varias resoluciones. En un grid de
    180px la versión 'small' pesa ~5 veces menos que 'normal'."""
    return (uri or "").replace("/normal/", "/small/")


def generate_html_coleccion(rows: list[dict], nombre: str, output_path,
                            sheet_for=None, oracle_completo: bool = True,
                            logo: str = "", banner: str = "",
                            titulo: str = "", subtitulo: str = "",
                            banner_modo: str = "hero"):
    # ── Preparar datos ────────────────────────────────────────────────────────
    cards, comandantes = [], {}
    sets, binders, condiciones, idiomas = defaultdict(int), defaultdict(int), set(), set()
    subtipo_count = defaultdict(int)
    color_count = defaultdict(int)

    for r in rows:
        tl   = r.get("type_line", "")
        subs = extraer_subtipos(tl)
        ci   = r.get("color_identity", "")
        qty  = r.get("qty", 1) if isinstance(r.get("qty"), int) else 1

        oracle = r.get("oracle_text", "") or ""
        if not oracle_completo and len(oracle) > 400:
            oracle = oracle[:400] + "…"

        # Legendarias que pueden ser comandante → alimentan el buscador local
        es_legend = "Legendary" in tl and "Creature" in tl
        puede_ser = es_legend or "can be your commander" in oracle.lower()
        if puede_ser and r.get("name") not in comandantes:
            comandantes[r["name"]] = ci

        if r.get("set_code"):
            sets[r["set_code"]] += qty
        if r.get("binder_name"):
            binders[r["binder_name"]] += qty
        if r.get("condition"):
            condiciones.add(r["condition"])
        if r.get("language"):
            idiomas.add(r["language"])
        for s in subs:
            subtipo_count[s] += qty
        color_count[len(ci)] += qty

        cards.append({
            "n":  r.get("name", ""),
            "t":  sheet_for(r) if sheet_for else "Other",
            "tl": tl,
            "mc": r.get("mana_cost", ""),
            "c":  r.get("cmc", ""),
            "o":  oracle,
            "r":  r.get("rarity", ""),
            "i":  img_small(r.get("image_uri", "")),
            "u":  r.get("scryfall_uri", ""),
            "f":  1 if r.get("foil") else 0,
            "q":  qty,
            "s":  r.get("set_code", ""),
            "sn": r.get("set_name", ""),
            "ci": ci,
            "co": r.get("colors", ""),
            "sb": subs,
            "cn": r.get("collector_number", ""),
            "cd": r.get("condition", ""),
            "lg": r.get("language", ""),
            "b":  r.get("binder_name", ""),
            "p":  r.get("power", ""),
            "tg": r.get("toughness", ""),
            "k":  r.get("keywords", ""),
        })

    total_distintas = len(cards)
    total_copias    = sum(c["q"] for c in cards)
    total_foils     = sum(c["q"] for c in cards if c["f"])
    total_sets      = len(sets)

    subtipos_ord = sorted(subtipo_count.items(), key=lambda kv: kv[0].lower())
    sets_ord     = sorted(sets.items())
    binders_ord  = sorted(binders.items())
    comandantes_ord = sorted(comandantes.items(), key=lambda kv: kv[0].lower())

    tipos_presentes = sorted({c["t"] for c in cards})

    opciones_subtipo = "".join(
        f'<option value="{s}">{s} ({n})</option>' for s, n in subtipos_ord)
    opciones_set = "".join(
        f'<option value="{s}">{s} ({n})</option>' for s, n in sets_ord)
    opciones_binder = "".join(
        f'<option value="{b}">{b} ({n})</option>' for b, n in binders_ord)
    opciones_tipo = "".join(
        f'<option value="{t}">{t}</option>' for t in tipos_presentes)
    opciones_cond = "".join(
        f'<option value="{c}">{c}</option>' for c in sorted(condiciones))
    opciones_idioma = "".join(
        f'<option value="{i}">{i}</option>' for i in sorted(idiomas))
    datalist_cmd = "".join(
        f'<option value="{n}">' for n, _ in comandantes_ord)

    html = PLANTILLA
    bloque_logo = f'<img class="logo" src="{logo}" alt="logo">' if logo else ""

    # Modo hero: el banner se muestra entero y el título va dentro de la imagen
    if banner and banner_modo == "hero":
        bloque_banner  = f'<img class="hero-banner" src="{banner}" alt="{titulo or nombre}">'
        estilo_banner  = "padding:1.25rem 1rem 1.75rem;"
        bloque_titulo  = ""
    elif banner:
        bloque_banner  = ""
        estilo_banner  = (
            f'background-image:linear-gradient(rgba(10,10,20,.72),rgba(10,10,20,.88)),url({banner});'
            f'background-size:cover;background-position:center;')
        bloque_titulo  = f"<h1>{titulo or nombre}</h1>"
    else:
        bloque_banner  = ""
        estilo_banner  = ""
        bloque_titulo  = f"<h1>{titulo or nombre}</h1>"

    reemplazos = {
        "__NOMBRE__":        nombre,
        "__TITULO__":        titulo or nombre,
        "__TITULO_HTML__":   bloque_titulo,
        "__BANNER_IMG__":    bloque_banner,
        "__SUBTITULO__":     subtitulo or "Colección de Magic: The Gathering · disponible para intercambio",
        "__LOGO_HTML__":     bloque_logo,
        "__BANNER_STYLE__":  estilo_banner,
        "__DISTINTAS__":     f"{total_distintas:,}".replace(",", "."),
        "__COPIAS__":        f"{total_copias:,}".replace(",", "."),
        "__FOILS__":         f"{total_foils:,}".replace(",", "."),
        "__SETS__":          str(total_sets),
        "__NSUBTIPOS__":     str(len(subtipos_ord)),
        "__NCOMANDANTES__":  str(len(comandantes_ord)),
        "__OPC_SUBTIPO__":   opciones_subtipo,
        "__OPC_SET__":       opciones_set,
        "__OPC_BINDER__":    opciones_binder,
        "__OPC_TIPO__":      opciones_tipo,
        "__OPC_COND__":      opciones_cond,
        "__OPC_IDIOMA__":    opciones_idioma,
        "__DATALIST_CMD__":  datalist_cmd,
        "__CARDS_JSON__":    json.dumps(cards, ensure_ascii=False, separators=(",", ":")),
        "__CMD_JSON__":      json.dumps(dict(comandantes_ord), ensure_ascii=False),
        "__BLOQUE_BINDER__": "" if not binders_ord else BLOQUE_BINDER,
        "__BLOQUE_COND__":   "" if not condiciones else BLOQUE_COND,
        "__BLOQUE_IDIOMA__": "" if not idiomas else BLOQUE_IDIOMA,
    }
    for token, valor in reemplazos.items():
        html = html.replace(token, valor)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)


BLOQUE_BINDER = """
        <div class="fgroup">
          <label>Carpeta</label>
          <select class="msel" id="f-binder"><option value="">Todas</option>__OPC_BINDER__</select>
        </div>"""

BLOQUE_COND = """
        <div class="fgroup">
          <label>Condición</label>
          <select class="msel" id="f-cond"><option value="">Todas</option>__OPC_COND__</select>
        </div>"""

BLOQUE_IDIOMA = """
        <div class="fgroup">
          <label>Idioma</label>
          <select class="msel" id="f-lang"><option value="">Todos</option>__OPC_IDIOMA__</select>
        </div>"""


PLANTILLA = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__NOMBRE__ — Colección MTG</title>
<style>
  :root {
    --bg:#0d0d0d; --surface:#1a1a2e; --surface2:#16213e; --accent:#e94560;
    --text:#e6e6e6; --text2:#9aa0aa; --border:#2a2a4a; --ok:#2d9d78;
    --card-w:170px;
  }
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:var(--bg);color:var(--text);font-family:'Segoe UI',Arial,sans-serif;padding-bottom:80px}

  header{background:linear-gradient(135deg,#1a1a2e,#16213e 60%,#0f3460);
         padding:2rem 1.5rem;border-bottom:1px solid var(--border);text-align:center;
         background-size:cover;background-position:center}
  header .logo{max-height:110px;max-width:min(320px,80vw);margin:0 auto 1rem;display:block;
               filter:drop-shadow(0 4px 14px rgba(0,0,0,.65))}
  header .hero-banner{width:100%;max-width:1100px;margin:0 auto .9rem;display:block;
                      border-radius:14px;box-shadow:0 10px 34px rgba(0,0,0,.6)}
  header h1{font-size:2rem;color:#fff;margin-bottom:.35rem;
            text-shadow:0 2px 12px rgba(0,0,0,.8)}
  header p{color:#c9cfd8;font-size:.9rem;text-shadow:0 2px 8px rgba(0,0,0,.8)}
  .stats{display:flex;gap:.75rem;justify-content:center;flex-wrap:wrap;margin-top:1.25rem}
  .stat{background:rgba(0,0,0,.3);border:1px solid var(--border);border-radius:10px;
        padding:.5rem 1.1rem;min-width:110px}
  .stat b{display:block;font-size:1.4rem;color:var(--accent)}
  .stat span{font-size:.7rem;color:var(--text2);text-transform:uppercase;letter-spacing:.05em}

  main{max-width:1500px;margin:0 auto;padding:1.5rem}

  /* ── Panel de filtros ── */
  .panel{background:var(--surface);border:1px solid var(--border);border-radius:12px;
         padding:1rem;margin-bottom:1rem}
  .frow{display:flex;flex-wrap:wrap;gap:1rem;align-items:flex-end}
  .frow + .frow{margin-top:.9rem;padding-top:.9rem;border-top:1px solid var(--border)}
  .fgroup{display:flex;flex-direction:column;gap:.3rem}
  .fgroup label{font-size:.68rem;text-transform:uppercase;letter-spacing:.07em;color:var(--text2)}
  .msel,.mtxt{background:var(--surface2);border:1px solid var(--border);color:var(--text);
              padding:.42rem .6rem;border-radius:8px;font-size:.8rem;outline:none}
  .msel:focus,.mtxt:focus{border-color:var(--accent)}
  .mtxt.wide{min-width:250px}
  .pips{display:flex;gap:.4rem;align-items:center}
  .pip{width:32px;height:32px;border-radius:50%;cursor:pointer;border:2px solid var(--border);
       display:flex;align-items:center;justify-content:center;font-weight:800;font-size:.8rem;
       user-select:none;transition:transform .12s}
  .pip:hover{transform:translateY(-2px)}
  .pip.on{border-color:#fff;box-shadow:0 0 0 2px var(--accent)}
  .pip-W{background:#f8f6d8;color:#6b6242}.pip-U{background:#c1d7e9;color:#1f4d68}
  .pip-B{background:#4a4a4a;color:#d9d0d0}.pip-R{background:#e4a08a;color:#7a2f1c}
  .pip-G{background:#a3c095;color:#23502a}.pip-C{background:#cac5c0;color:#4a4a4a}
  .chip{background:var(--surface2);border:1px solid var(--border);color:var(--text2);
        padding:.35rem .8rem;border-radius:14px;cursor:pointer;font-size:.75rem}
  .chip.on{border-color:var(--accent);background:var(--accent);color:#fff}
  .btn{background:var(--surface2);border:1px solid var(--accent);color:var(--text);
       padding:.42rem 1rem;border-radius:8px;cursor:pointer;font-size:.78rem}
  .btn:hover{background:var(--accent);color:#fff}
  .btn.solid{background:var(--accent);color:#fff}
  #cmd-status{font-size:.75rem;color:var(--ok);min-height:1em}

  .bar{display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;
       gap:.75rem;margin-bottom:1rem}
  .count{font-size:.82rem;color:var(--text2)}
  .count b{color:var(--accent)}

  /* ── Galería ── */
  #grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(var(--card-w),1fr));gap:.9rem}

  /* En móvil una carta por fila desperdicia pantalla y obliga a estirar la
     imagen. Bajamos el ancho mínimo para que entren 2-3 por fila. */
  @media (max-width:900px){
    :root{--card-w:140px}
    main{padding:1rem}
    header{padding:1.25rem .75rem}
    header h1{font-size:1.5rem}
    .panel{padding:.75rem}
    #grid{gap:.6rem}
  }
  @media (max-width:600px){
    :root{--card-w:120px}
    main{padding:.6rem}
    .frow{gap:.6rem}
    .mtxt.wide{min-width:100%}
    .fgroup{flex:1 1 100%}
    .fgroup .pips{justify-content:space-between;width:100%}
    .msel,.mtxt{width:100%;max-width:100%}
    .cfoot{font-size:.66rem;padding:.35rem .4rem}
    .stat{padding:.4rem .7rem;min-width:88px}
    .stat b{font-size:1.1rem}
    table{font-size:.72rem}
  }
  .card{background:var(--surface);border:1px solid var(--border);border-radius:10px;
        overflow:hidden;position:relative;cursor:pointer;transition:transform .15s,border-color .15s}
  .card:hover{transform:translateY(-4px);border-color:var(--accent)}
  .card.picked{border-color:var(--ok);box-shadow:0 0 0 2px var(--ok)}
  .card img{width:100%;display:block;aspect-ratio:488/680;object-fit:cover;background:#111}
  .noimg{aspect-ratio:488/680;display:flex;flex-direction:column;align-items:center;
         justify-content:center;padding:.5rem;text-align:center;font-size:.7rem;color:var(--text2)}
  .cfoot{padding:.45rem .55rem;font-size:.72rem}
  .cfoot .nm{font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  .cfoot .mt{color:var(--text2);font-size:.66rem;margin-top:2px}
  .badge{position:absolute;top:5px;right:5px;background:rgba(0,0,0,.75);border-radius:6px;
         padding:1px 5px;font-size:.62rem;z-index:2}
  .badge.foil{color:#ffd700}
  .pick{position:absolute;top:5px;left:5px;width:26px;height:26px;border-radius:6px;
        background:rgba(0,0,0,.7);border:1px solid var(--border);color:#fff;font-size:1rem;
        display:flex;align-items:center;justify-content:center;z-index:3}
  .card.picked .pick{background:var(--ok);border-color:var(--ok)}

  /* ── Tabla ── */
  table{width:100%;border-collapse:collapse;font-size:.8rem}
  th{background:var(--surface2);padding:.55rem .5rem;text-align:left;position:sticky;top:0;
     font-size:.7rem;text-transform:uppercase;letter-spacing:.05em;color:var(--text2);z-index:5}
  td{padding:.45rem .5rem;border-bottom:1px solid var(--border)}
  tr:hover td{background:var(--surface)}
  tr.picked td{background:rgba(45,157,120,.12)}
  .tname{color:var(--text);font-weight:600;cursor:pointer}
  .tname:hover{color:var(--accent)}

  /* ── Barra de intercambio ── */
  #tradebar{position:fixed;bottom:0;left:0;right:0;background:var(--surface);
            border-top:2px solid var(--ok);padding:.7rem 1.5rem;display:none;
            justify-content:space-between;align-items:center;gap:1rem;flex-wrap:wrap;z-index:200}
  #tradebar.show{display:flex}
  #tradebar .info{font-size:.85rem}
  #tradebar .info b{color:var(--ok)}

  /* ── Modal ── */
  .overlay{position:fixed;inset:0;background:rgba(0,0,0,.85);display:none;
           align-items:center;justify-content:center;z-index:300;padding:1rem}
  .overlay.show{display:flex}
  .modal{background:var(--surface);border:1px solid var(--border);border-radius:14px;
         max-width:980px;width:100%;max-height:92vh;overflow-y:auto;padding:1.5rem;
         display:flex;gap:1.75rem;flex-wrap:wrap;position:relative}
  .modal img{flex:0 0 380px;width:380px;max-width:100%;border-radius:14px;
             box-shadow:0 8px 30px rgba(0,0,0,.6)}
  @media (max-width:760px){
    .modal{flex-direction:column;padding:1rem}
    .modal img{flex:none;width:100%;max-width:420px;margin:0 auto}
  }
  .modal .info{flex:1;min-width:260px}
  .modal h2{color:#fff;margin-bottom:.5rem}
  .modal .ot{white-space:pre-wrap;background:var(--surface2);padding:.8rem;border-radius:8px;
             font-size:.85rem;line-height:1.5;margin:.8rem 0}
  .kv{display:flex;justify-content:space-between;padding:.28rem 0;
      border-bottom:1px solid var(--border);font-size:.8rem}
  .kv span:first-child{color:var(--text2)}
  .close{position:absolute;top:1rem;right:1.5rem;font-size:2rem;color:var(--text2);cursor:pointer}
  .empty{text-align:center;color:var(--text2);padding:3rem 1rem}
  .more{display:block;margin:2rem auto 0;background:var(--surface2);border:1px solid var(--accent);
        color:var(--text);padding:.7rem 2rem;border-radius:20px;cursor:pointer}
  textarea{width:100%;height:220px;background:var(--bg);color:var(--text);border:1px solid var(--border);
           border-radius:8px;padding:.7rem;font-family:Consolas,monospace;font-size:.8rem}
</style>
</head>
<body>

<header style="__BANNER_STYLE__">
  __BANNER_IMG__
  __LOGO_HTML__
  __TITULO_HTML__
  <p>__SUBTITULO__</p>
  <div class="stats">
    <div class="stat"><b>__DISTINTAS__</b><span>Cartas distintas</span></div>
    <div class="stat"><b>__COPIAS__</b><span>Copias totales</span></div>
    <div class="stat"><b>__SETS__</b><span>Ediciones</span></div>
    <div class="stat"><b>__FOILS__</b><span>Foils</span></div>
    <div class="stat"><b>__NSUBTIPOS__</b><span>Subtipos</span></div>
  </div>
</header>

<main>
  <div class="panel">

    <!-- FILA 1: comandante + búsqueda -->
    <div class="frow">
      <div class="fgroup" style="flex:1;min-width:280px">
        <label>Identidad de comandante — muestra solo lo legal en su mazo</label>
        <div style="display:flex;gap:.5rem;flex-wrap:wrap">
          <input class="mtxt wide" id="cmd-input" list="cmd-list"
                 placeholder="Escribe un comandante: Gishath, Atraxa, Muldrotha...">
          <datalist id="cmd-list">__DATALIST_CMD__</datalist>
          <button class="btn solid" id="cmd-go">Aplicar</button>
          <button class="btn" id="cmd-clear">Quitar</button>
        </div>
        <div id="cmd-status">__NCOMANDANTES__ comandantes detectados en esta colección</div>
      </div>
      <div class="fgroup" style="flex:1;min-width:240px">
        <label>Buscar por nombre o texto de la carta</label>
        <input class="mtxt wide" id="f-text" placeholder="Ej: sacrifice, Sol Ring, draw a card...">
      </div>
    </div>

    <!-- FILA 2: colores -->
    <div class="frow">
      <div class="fgroup">
        <label>Color</label>
        <div class="pips">
          <div class="pip pip-W" data-c="W" title="Blanco">W</div>
          <div class="pip pip-U" data-c="U" title="Azul">U</div>
          <div class="pip pip-B" data-c="B" title="Negro">B</div>
          <div class="pip pip-R" data-c="R" title="Rojo">R</div>
          <div class="pip pip-G" data-c="G" title="Verde">G</div>
          <div class="pip pip-C" data-c="C" title="Incoloro">C</div>
        </div>
      </div>
      <div class="fgroup">
        <label>Modo</label>
        <select class="msel" id="f-cmode">
          <option value="subset">Legal en esa identidad</option>
          <option value="any">Contiene alguno</option>
          <option value="all">Contiene todos</option>
          <option value="exact">Exactamente esos</option>
        </select>
      </div>
      <div class="fgroup">
        <label>Cantidad de colores</label>
        <div style="display:flex;gap:.4rem">
          <button class="chip on" data-cc="todos">Todas</button>
          <button class="chip" data-cc="mono">Mono</button>
          <button class="chip" data-cc="multi">Multi</button>
          <button class="chip" data-cc="incoloro">Incoloras</button>
        </div>
      </div>
    </div>

    <!-- FILA 3: tipo, subtipo, rareza, etc -->
    <div class="frow">
      <div class="fgroup">
        <label>Tipo</label>
        <select class="msel" id="f-tipo"><option value="">Todos</option>__OPC_TIPO__</select>
      </div>
      <div class="fgroup">
        <label>Subtipo (__NSUBTIPOS__)</label>
        <select class="msel" id="f-sub" style="max-width:200px">
          <option value="">Todos</option>__OPC_SUBTIPO__
        </select>
      </div>
      <div class="fgroup">
        <label>o escribe el subtipo</label>
        <input class="mtxt" id="f-subtxt" placeholder="elf, snake, dragon..." style="width:150px">
      </div>
      <div class="fgroup">
        <label>Rareza</label>
        <select class="msel" id="f-rar">
          <option value="">Todas</option>
          <option value="mythic">Mythic</option><option value="rare">Rare</option>
          <option value="uncommon">Uncommon</option><option value="common">Common</option>
        </select>
      </div>
      <div class="fgroup">
        <label>Edición</label>
        <select class="msel" id="f-set" style="max-width:170px"><option value="">Todas</option>__OPC_SET__</select>
      </div>
__BLOQUE_BINDER__
__BLOQUE_COND__
__BLOQUE_IDIOMA__
      <div class="fgroup">
        <label>CMC</label>
        <div style="display:flex;gap:.3rem;align-items:center">
          <input class="mtxt" id="f-cmin" type="number" min="0" max="20" placeholder="min" style="width:65px">
          <span style="color:var(--text2)">–</span>
          <input class="mtxt" id="f-cmax" type="number" min="0" max="20" placeholder="max" style="width:65px">
        </div>
      </div>
      <div class="fgroup">
        <label>&nbsp;</label>
        <div style="display:flex;gap:.4rem">
          <button class="chip" id="f-foil">Solo foil</button>
          <button class="btn" id="f-clear">Limpiar todo</button>
        </div>
      </div>
    </div>
  </div>

  <div class="bar">
    <div class="count" id="count"></div>
    <div style="display:flex;gap:.5rem;flex-wrap:wrap">
      <select class="msel" id="f-sort">
        <option value="name">Nombre A-Z</option>
        <option value="cmc">CMC</option>
        <option value="rarity">Rareza</option>
        <option value="set">Edición</option>
        <option value="qty">Cantidad</option>
      </select>
      <select class="msel" id="f-size" title="Tamaño de las cartas">
        <option value="">Tamaño automático</option>
        <option value="110">Cartas pequeñas</option>
        <option value="170">Cartas medianas</option>
        <option value="240">Cartas grandes</option>
      </select>
      <button class="chip on" id="v-grid">Galería</button>
      <button class="chip" id="v-table">Tabla</button>
      <button class="btn" id="export-filtro">Exportar lo filtrado</button>
    </div>
  </div>

  <div id="grid"></div>
  <div id="tablewrap" style="display:none"></div>
</main>

<!-- Barra de intercambio -->
<div id="tradebar">
  <div class="info">Lista de intercambio: <b id="pick-count">0</b> cartas seleccionadas</div>
  <div style="display:flex;gap:.5rem;flex-wrap:wrap">
    <button class="btn" id="pick-view">Ver lista</button>
    <button class="btn solid" id="pick-copy">Copiar</button>
    <button class="btn" id="pick-download">Descargar TXT</button>
    <button class="btn" id="pick-clear">Vaciar</button>
  </div>
</div>

<!-- Modal -->
<div class="overlay" id="overlay">
  <div class="modal" id="modal"></div>
</div>

<script>
const CARDS = __CARDS_JSON__;
CARDS.forEach((c,i) => c._i = i);
const CMDS  = __CMD_JSON__;
const PAGE  = 200;
const RCOL  = {common:"#aaa",uncommon:"#8ab4f8",rare:"#ffd700",mythic:"#ff8c00",special:"#9b59b6"};

let shown = PAGE, view = "grid";
let colors = new Set(), cmode = "subset", ccount = "todos", onlyFoil = false;
let picked = new Set();
let filtered = [];

const $ = id => document.getElementById(id);
const val = id => { const e = $(id); return e ? e.value : ""; };

/* ── Identidad de color ────────────────────────────────────────────────────
   subset = la identidad de la carta cabe dentro de la seleccionada.
   Es la regla real de Commander: una carta es legal si su identidad de color
   es un subconjunto de la del comandante. Las incoloras siempre son legales. */
function colorOK(c) {
  const arr = (c.ci || "").split("").filter(Boolean);
  const n = arr.length;
  if (ccount === "mono" && n !== 1) return false;
  if (ccount === "multi" && n < 2) return false;
  if (ccount === "incoloro" && n !== 0) return false;
  if (!colors.size) return true;

  const sel = [...colors], conC = sel.includes("C"), cols = sel.filter(x => x !== "C");
  if (cmode === "subset") return arr.every(x => cols.includes(x));
  if (cmode === "any")    return (conC && n === 0) || cols.some(x => arr.includes(x));
  if (cmode === "all")    return cols.length ? cols.every(x => arr.includes(x)) : (conC ? n === 0 : true);
  if (!cols.length) return n === 0;
  return n === cols.length && cols.every(x => arr.includes(x));
}

function aplicarFiltros() {
  const q    = val("f-text").toLowerCase().trim();
  const tipo = val("f-tipo");
  let   sub  = val("f-sub") || val("f-subtxt").trim();
  const rar  = val("f-rar"), set = val("f-set");
  const bind = val("f-binder"), cond = val("f-cond"), lang = val("f-lang");
  const cmin = val("f-cmin") === "" ? -1 : +val("f-cmin");
  const cmax = val("f-cmax") === "" ? 99 : +val("f-cmax");
  const subq = sub.toLowerCase();

  return CARDS.filter(c => {
    if (q && !c.n.toLowerCase().includes(q) && !(c.o || "").toLowerCase().includes(q)) return false;
    if (tipo && c.t !== tipo) return false;
    if (sub && !(c.sb || []).some(s => s.toLowerCase().includes(subq))) return false;
    if (rar && c.r !== rar) return false;
    if (set && c.s !== set) return false;
    if (bind && c.b !== bind) return false;
    if (cond && c.cd !== cond) return false;
    if (lang && c.lg !== lang) return false;
    if (onlyFoil && !c.f) return false;
    if (c.c !== "" && (c.c < cmin || c.c > cmax)) return false;
    return colorOK(c);
  });
}

const SORTS = {
  name: (a,b) => a.n.localeCompare(b.n),
  cmc:  (a,b) => (a.c === "" ? 99 : a.c) - (b.c === "" ? 99 : b.c) || a.n.localeCompare(b.n),
  rarity:(a,b)=> ({mythic:0,rare:1,uncommon:2,common:3}[a.r] ?? 9) - ({mythic:0,rare:1,uncommon:2,common:3}[b.r] ?? 9) || a.n.localeCompare(b.n),
  set:  (a,b) => (a.s||"").localeCompare(b.s||"") || a.n.localeCompare(b.n),
  qty:  (a,b) => b.q - a.q || a.n.localeCompare(b.n),
};

function render(reset) {
  if (reset !== false) shown = PAGE;
  filtered = aplicarFiltros();
  filtered.sort(SORTS[val("f-sort")] || SORTS.name);

  const copias = filtered.reduce((s,c) => s + c.q, 0);
  $("count").innerHTML = `<b>${filtered.length}</b> cartas distintas · <b>${copias}</b> copias` +
    (filtered.length > shown ? ` · mostrando ${shown}` : "");

  const vis = filtered.slice(0, shown);
  const grid = $("grid"), tw = $("tablewrap");

  if (!filtered.length) {
    grid.style.display = "block"; tw.style.display = "none";
    grid.innerHTML = '<div class="empty">Ninguna carta coincide con estos filtros.</div>';
    return;
  }

  if (view === "grid") {
    grid.style.display = "grid"; tw.style.display = "none";
    grid.innerHTML = vis.map(cardHTML).join("");
    if (filtered.length > shown) grid.appendChild(botonMas());
  } else {
    grid.style.display = "none"; tw.style.display = "block";
    tw.innerHTML = tablaHTML(vis);
    if (filtered.length > shown) tw.appendChild(botonMas());
  }
}

function botonMas() {
  const b = document.createElement("button");
  b.className = "more";
  b.textContent = `Cargar ${Math.min(PAGE, filtered.length - shown)} más (quedan ${filtered.length - shown})`;
  b.onclick = () => { shown += PAGE; render(false); };
  return b;
}

function idx(c) { return c._i; }

/* Scryfall sirve la misma imagen en varias resoluciones:
     small 146px · normal 488px · large 672px
   Con srcset el navegador elige según el ancho real de la tarjeta Y la densidad
   de la pantalla. En un móvil 3x una tarjeta de 160 CSS px son 480 px reales,
   así que baja la 'normal' en vez de estirar la 'small' y verse borrosa. */
const IMG_SIZES = "(max-width:600px) 46vw, (max-width:900px) 24vw, 210px";

function srcsetDe(u) {
  if (!u) return "";
  return `${u} 146w, ${u.replace("/small/","/normal/")} 488w, ${u.replace("/small/","/large/")} 672w`;
}

function cardHTML(c) {
  const i = idx(c);
  const img = c.i
    ? `<img src="${c.i.replace("/small/","/normal/")}" srcset="${srcsetDe(c.i)}"
            sizes="${IMG_SIZES}" alt="${c.n}" loading="lazy" decoding="async">`
    : `<div class="noimg"><div style="font-size:1.6rem">🃏</div><div>${c.n}</div><div>${c.tl}</div></div>`;
  return `<div class="card ${picked.has(i) ? "picked" : ""}" data-i="${i}">
    <div class="pick" onclick="togglePick(event,${i})">${picked.has(i) ? "✓" : "+"}</div>
    ${c.f ? '<div class="badge foil">✦</div>' : ""}
    ${c.q > 1 ? `<div class="badge" style="top:auto;bottom:44px">×${c.q}</div>` : ""}
    <div onclick="abrir(${i})">${img}</div>
    <div class="cfoot" onclick="abrir(${i})">
      <div class="nm">${c.n}</div>
      <div class="mt">${c.mc || ""} ${c.s ? "· " + c.s : ""}</div>
    </div>
  </div>`;
}

function tablaHTML(vis) {
  const filas = vis.map(c => {
    const i = idx(c);
    return `<tr class="${picked.has(i) ? "picked" : ""}">
      <td><button class="chip ${picked.has(i) ? "on" : ""}" onclick="togglePick(event,${i})">${picked.has(i) ? "✓" : "+"}</button></td>
      <td class="tname" onclick="abrir(${i})">${c.n}${c.f ? ' <span style="color:#ffd700">✦</span>' : ""}</td>
      <td>${c.q}</td><td>${c.mc || ""}</td><td>${c.c === "" ? "" : c.c}</td>
      <td style="font-size:.72rem">${c.tl}</td>
      <td><span style="color:${RCOL[c.r] || "#888"}">●</span> ${c.s}</td>
      <td>${c.cn || ""}</td><td>${c.cd || ""}</td><td>${c.b || ""}</td>
    </tr>`;
  }).join("");
  return `<table><thead><tr>
    <th></th><th>Carta</th><th>Cant</th><th>Coste</th><th>CMC</th>
    <th>Tipo</th><th>Edición</th><th>N.º</th><th>Cond</th><th>Carpeta</th>
  </tr></thead><tbody>${filas}</tbody></table>`;
}

/* ── Selección para intercambio ── */
function togglePick(ev, i) {
  ev.stopPropagation();
  picked.has(i) ? picked.delete(i) : picked.add(i);
  $("pick-count").textContent = picked.size;
  $("tradebar").classList.toggle("show", picked.size > 0);
  render(false);
}

function listaTexto(items) {
  return items.map(c => `${c.q} ${c.n}${c.s ? ` (${c.s})` : ""}${c.cn ? ` ${c.cn}` : ""}${c.f ? " *F*" : ""}`).join("\n");
}

function descargar(texto, nombre) {
  const a = document.createElement("a");
  a.href = URL.createObjectURL(new Blob([texto], {type:"text/plain"}));
  a.download = nombre; a.click(); URL.revokeObjectURL(a.href);
}

function copiar(texto, btn) {
  navigator.clipboard.writeText(texto).then(() => {
    const t = btn.textContent; btn.textContent = "¡Copiado!";
    setTimeout(() => btn.textContent = t, 1500);
  }).catch(() => alert("Tu navegador bloqueó el portapapeles. Usa 'Ver lista' y copia a mano."));
}

/* ── Modal ── */
function abrir(i) {
  const c = CARDS[i];
  const kv = (k,v) => v ? `<div class="kv"><span>${k}</span><span>${v}</span></div>` : "";
  $("modal").innerHTML = `
    <span class="close" onclick="cerrar()">&times;</span>
    ${c.i ? `<img src="${c.i.replace("/small/","/large/")}" alt="${c.n}" loading="eager">` : ""}
    <div class="info">
      <h2>${c.n}</h2>
      <div style="color:var(--text2);font-size:.85rem">${c.tl}</div>
      ${c.o ? `<div class="ot">${c.o}</div>` : ""}
      ${kv("Coste", c.mc)}${kv("CMC", c.c)}${kv("Fuerza/Resistencia", c.p ? c.p + "/" + c.tg : "")}
      ${kv("Identidad de color", c.ci || "Incolora")}${kv("Rareza", c.r)}
      ${kv("Edición", (c.sn || c.s) + (c.cn ? " · #" + c.cn : ""))}
      ${kv("Copias", c.q)}${kv("Foil", c.f ? "Sí" : "")}${kv("Condición", c.cd)}
      ${kv("Idioma", c.lg)}${kv("Carpeta", c.b)}
      <div style="margin-top:1rem;display:flex;gap:.5rem;flex-wrap:wrap">
        <button class="btn solid" onclick="togglePick(event,${i})">
          ${picked.has(i) ? "Quitar de la lista" : "Añadir a intercambio"}</button>
        ${c.u ? `<a class="btn" href="${c.u}" target="_blank">Ver en Scryfall</a>` : ""}
      </div>
    </div>`;
  $("overlay").classList.add("show");
}
function cerrar() { $("overlay").classList.remove("show"); }

/* ── Comandante ── */
function aplicarIdentidad(nombre, ci) {
  colors.clear();
  (ci || "").split("").forEach(x => colors.add(x));
  // Comandante incoloro (Kozilek, Karn...): marcamos C para que en modo subset
  // solo pasen las cartas sin identidad de color, que es lo legal en su mazo
  if (!ci) colors.add("C");
  document.querySelectorAll(".pip").forEach(p => p.classList.toggle("on", colors.has(p.dataset.c)));
  cmode = "subset"; $("f-cmode").value = "subset";
  $("cmd-status").textContent = ci
    ? `Mostrando lo legal en ${nombre} — identidad ${ci}`
    : `${nombre} es incoloro — solo cartas incoloras`;
  render();
}

async function buscarComandante() {
  const nombre = $("cmd-input").value.trim();
  if (!nombre) return;
  // 1) Buscar en la propia colección (instantáneo y funciona sin internet)
  const exacto = Object.keys(CMDS).find(n => n.toLowerCase() === nombre.toLowerCase());
  const parcial = exacto || Object.keys(CMDS).find(n => n.toLowerCase().includes(nombre.toLowerCase()));
  if (parcial) { aplicarIdentidad(parcial, CMDS[parcial]); return; }
  // 2) Si no está, preguntarle a Scryfall
  $("cmd-status").textContent = "Buscando en Scryfall...";
  try {
    const r = await fetch("https://api.scryfall.com/cards/named?fuzzy=" + encodeURIComponent(nombre));
    if (!r.ok) throw new Error("no encontrado");
    const d = await r.json();
    aplicarIdentidad(d.name, (d.color_identity || []).join(""));
  } catch {
    $("cmd-status").textContent = "No se encontró ese comandante.";
  }
}

/* ── Eventos ── */
["f-text","f-subtxt","f-cmin","f-cmax"].forEach(id => $(id).addEventListener("input", () => render()));
["f-tipo","f-sub","f-rar","f-set","f-sort","f-binder","f-cond","f-lang"].forEach(id => {
  const e = $(id); if (e) e.addEventListener("change", () => render());
});
$("f-sub").addEventListener("change", () => { if (val("f-sub")) $("f-subtxt").value = ""; });
$("f-subtxt").addEventListener("input", () => { if (val("f-subtxt")) $("f-sub").value = ""; });

document.querySelectorAll(".pip").forEach(p => p.addEventListener("click", () => {
  const c = p.dataset.c;
  colors.has(c) ? colors.delete(c) : colors.add(c);
  p.classList.toggle("on", colors.has(c));
  render();
}));
$("f-cmode").addEventListener("change", e => { cmode = e.target.value; render(); });
document.querySelectorAll("[data-cc]").forEach(ch => ch.addEventListener("click", () => {
  document.querySelectorAll("[data-cc]").forEach(x => x.classList.remove("on"));
  ch.classList.add("on"); ccount = ch.dataset.cc; render();
}));
$("f-foil").addEventListener("click", () => {
  onlyFoil = !onlyFoil; $("f-foil").classList.toggle("on", onlyFoil); render();
});
$("v-grid").addEventListener("click", () => {
  view = "grid"; $("v-grid").classList.add("on"); $("v-table").classList.remove("on"); render(false);
});
$("f-size").addEventListener("change", e => {
  // Vacío = deja mandar a las media queries del CSS
  if (e.target.value) document.documentElement.style.setProperty("--card-w", e.target.value + "px");
  else document.documentElement.style.removeProperty("--card-w");
});
$("v-table").addEventListener("click", () => {
  view = "table"; $("v-table").classList.add("on"); $("v-grid").classList.remove("on"); render(false);
});
$("cmd-go").addEventListener("click", buscarComandante);
$("cmd-input").addEventListener("keydown", e => { if (e.key === "Enter") buscarComandante(); });
$("cmd-clear").addEventListener("click", () => {
  colors.clear();
  document.querySelectorAll(".pip").forEach(p => p.classList.remove("on"));
  $("cmd-input").value = "";
  $("cmd-status").textContent = "__NCOMANDANTES__ comandantes detectados en esta colección";
  render();
});
$("f-clear").addEventListener("click", () => {
  colors.clear(); document.querySelectorAll(".pip").forEach(p => p.classList.remove("on"));
  cmode = "subset"; $("f-cmode").value = "subset";
  ccount = "todos";
  document.querySelectorAll("[data-cc]").forEach(x => x.classList.remove("on"));
  document.querySelector('[data-cc="todos"]').classList.add("on");
  onlyFoil = false; $("f-foil").classList.remove("on");
  ["f-text","f-subtxt","f-cmin","f-cmax","f-tipo","f-sub","f-rar","f-set","f-binder","f-cond","f-lang","f-cmd"]
    .forEach(id => { const e = $(id); if (e) e.value = ""; });
  $("cmd-input").value = "";
  $("cmd-status").textContent = "__NCOMANDANTES__ comandantes detectados en esta colección";
  render();
});

$("export-filtro").addEventListener("click", () => {
  if (!filtered.length) return;
  descargar(listaTexto(filtered), "__NOMBRE___filtrado.txt");
});
$("pick-copy").addEventListener("click", e => copiar(listaTexto([...picked].map(i => CARDS[i])), e.target));
$("pick-download").addEventListener("click", () => descargar(listaTexto([...picked].map(i => CARDS[i])), "intercambio.txt"));
$("pick-clear").addEventListener("click", () => {
  picked.clear(); $("pick-count").textContent = "0";
  $("tradebar").classList.remove("show"); render(false);
});
$("pick-view").addEventListener("click", () => {
  const txt = listaTexto([...picked].map(i => CARDS[i]));
  $("modal").innerHTML = `<span class="close" onclick="cerrar()">&times;</span>
    <div class="info" style="flex:1">
      <h2>Lista de intercambio (${picked.size})</h2>
      <p style="color:var(--text2);font-size:.8rem;margin-bottom:.7rem">
        Copia este texto y mándaselo al dueño de la colección.</p>
      <textarea readonly>${txt}</textarea>
    </div>`;
  $("overlay").classList.add("show");
});

$("overlay").addEventListener("click", e => { if (e.target.id === "overlay") cerrar(); });
document.addEventListener("keydown", e => { if (e.key === "Escape") cerrar(); });

render();
</script>
</body>
</html>
"""




# ============================================================================
# EXPORTADOR WEB — JSON para el frontend estático
# ============================================================================

SCHEMA_VERSION = 1
IMG_BASE = "https://cards.scryfall.io/"

SUBTIPOS_COMPUESTOS = ["Time Lord"]

COLS = ["id", "n", "tl", "mc", "cmc", "r", "s", "cn",
        "ci", "co", "f", "q", "b", "cd", "lg", "sb", "t", "pt"]


def extraer_subtipos(type_line: str) -> list[str]:
    subs = []
    for cara in (type_line or "").split("//"):
        if "—" not in cara:
            continue
        cola = cara.split("—", 1)[1].strip()
        for compuesto in SUBTIPOS_COMPUESTOS:
            if compuesto in cola:
                subs.append(compuesto)
                cola = cola.replace(compuesto, " ")
        subs.extend(p.strip() for p in cola.split() if p.strip())
    vistos, out = set(), []
    for s in subs:
        if s not in vistos:
            vistos.add(s)
            out.append(s)
    return out


def url_derivada(card_id: str, size: str = "small") -> str:
    """
    Scryfall sirve las imágenes en un patrón predecible:
        https://cards.scryfall.io/{size}/front/{id[0]}/{id[1]}/{id}.jpg
    Guardando solo el UUID ahorramos ~90 bytes por carta frente a almacenar
    dos URLs completas. El generador verifica el patrón y, si alguna carta no
    encaja, guarda su URL literal en 'overrides'.
    """
    if not card_id or len(card_id) < 2:
        return ""
    return f"{IMG_BASE}{size}/front/{card_id[0]}/{card_id[1]}/{card_id}.jpg"


class Indice:
    """Diccionario que convierte valores repetidos en índices enteros."""

    def __init__(self):
        self.valores = []
        self._pos = {}

    def idx(self, valor):
        if valor in (None, ""):
            return -1
        if valor not in self._pos:
            self._pos[valor] = len(self.valores)
            self.valores.append(valor)
        return self._pos[valor]


def exportar_web(rows: list[dict], destino: Path, nombre: str,
                 sheet_for=None, titulo: str = "", subtitulo: str = "",
                 banner: str = "", logo: str = "") -> dict:
    destino = Path(destino)
    data_dir = destino / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    i_tl   = Indice()   # type_line completas
    i_set  = Indice()   # códigos de edición
    i_bind = Indice()   # carpetas
    i_cond = Indice()   # condiciones
    i_lang = Indice()   # idiomas
    i_tipo = Indice()   # hoja/tipo principal
    i_sub  = Indice()   # subtipos

    RAREZAS = ["common", "uncommon", "rare", "mythic", "special", "bonus"]
    rar_pos = {r: i for i, r in enumerate(RAREZAS)}

    set_names   = {}
    sub_count   = defaultdict(int)
    set_count   = defaultdict(int)
    bind_count  = defaultdict(int)
    comandantes = {}
    oracle      = {}
    overrides   = {}

    filas = []
    total_copias = total_foils = 0
    sin_imagen = 0

    for row in rows:
        qty = row.get("qty", 1)
        qty = qty if isinstance(qty, int) else 1
        nombre_carta = row.get("name", "")
        tl  = row.get("type_line", "")
        ci  = row.get("color_identity", "")
        cid = (row.get("scryfall_id") or "").strip()
        img = row.get("image_uri", "")
        txt = row.get("oracle_text", "") or ""

        # Scryfall añade un parámetro de caché (?1783907348) que no aporta nada
        # y estorba tanto para comparar como para extraer el UUID
        img = img.split("?", 1)[0]

        # Si no tenemos el UUID lo recuperamos de la propia URL de imagen
        if not cid and img:
            trozo = img.rsplit("/", 1)[-1]
            if trozo.endswith(".jpg"):
                cid = trozo[:-4]

        subs = extraer_subtipos(tl)
        for s in subs:
            sub_count[s] += qty

        set_code = row.get("set_code", "")
        if set_code:
            set_count[set_code] += qty
            if row.get("set_name"):
                set_names.setdefault(set_code, row["set_name"])

        binder = row.get("binder_name", "")
        if binder:
            bind_count[binder] += qty

        # Comandantes potenciales, para el filtro de identidad
        if ("Legendary" in tl and "Creature" in tl) or "can be your commander" in txt.lower():
            comandantes.setdefault(nombre_carta, ci)

        if txt and nombre_carta not in oracle:
            oracle[nombre_carta] = txt

        # ¿La URL real coincide con el patrón derivable?
        idx_fila = len(filas)
        if img:
            # Comparamos solo la parte estable de la ruta: el tamaño (small,
            # normal, large) lo elige el navegador, así que no debe influir
            cola = f"/front/{cid[0]}/{cid[1]}/{cid}.jpg" if len(cid) >= 2 else None
            if not cola or not img.endswith(cola):
                overrides[str(idx_fila)] = img
        elif not cid:
            sin_imagen += 1

        power, tough = row.get("power", ""), row.get("toughness", "")
        pt = f"{power}/{tough}" if power != "" or tough != "" else ""

        cmc = row.get("cmc", "")
        cmc = cmc if isinstance(cmc, (int, float)) else None

        filas.append([
            cid,
            nombre_carta,
            i_tl.idx(tl),
            row.get("mana_cost", ""),
            cmc,
            rar_pos.get(row.get("rarity", ""), -1),
            i_set.idx(set_code),
            row.get("collector_number", ""),
            ci,
            row.get("colors", ""),
            1 if row.get("foil") else 0,
            qty,
            i_bind.idx(binder),
            i_cond.idx(row.get("condition", "")),
            i_lang.idx(row.get("language", "")),
            [i_sub.idx(s) for s in subs],
            i_tipo.idx(sheet_for(row) if sheet_for else "Other"),
            pt,
        ])

        total_copias += qty
        if row.get("foil"):
            total_foils += qty

    meta = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "nombre": nombre,
        "titulo": titulo or nombre,
        "subtitulo": subtitulo or "Colección de Magic: The Gathering · disponible para intercambio",
        "banner": banner,
        "logo": logo,
        "img_base": IMG_BASE,
        "totales": {
            "distintas": len(filas),
            "copias": total_copias,
            "foils": total_foils,
            "sets": len(set_count),
            "subtipos": len(i_sub.valores),
            "comandantes": len(comandantes),
        },
        "rarezas": RAREZAS,
        "tipos": i_tipo.valores,
        "typelines": i_tl.valores,
        "sets": [[c, set_names.get(c, ""), set_count.get(c, 0)] for c in i_set.valores],
        "binders": [[b, bind_count.get(b, 0)] for b in i_bind.valores],
        "condiciones": i_cond.valores,
        "idiomas": i_lang.valores,
        "subtipos": [[s, sub_count.get(s, 0)] for s in i_sub.valores],
        "comandantes": dict(sorted(comandantes.items(), key=lambda kv: kv[0].lower())),
    }

    cards = {
        "schema_version": SCHEMA_VERSION,
        "cols": COLS,
        "rows": filas,
        "overrides": overrides,
    }

    oracle_doc = {"schema_version": SCHEMA_VERSION, "oracle": oracle}

    def escribir(nombre_archivo, doc):
        ruta = data_dir / nombre_archivo
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(doc, f, ensure_ascii=False, separators=(",", ":"))
        return ruta.stat().st_size

    tam = {
        "meta.json":   escribir("meta.json", meta),
        "cards.json":  escribir("cards.json", cards),
        "oracle.json": escribir("oracle.json", oracle_doc),
    }

    print(f"\n  Datos web escritos en {data_dir}")
    for archivo, bytes_ in tam.items():
        print(f"    {archivo:<14} {bytes_/1024:>8.0f} KB")
    if overrides:
        print(f"    (URLs de imagen no derivables: {len(overrides)})")
    if sin_imagen:
        print(f"    ⚠ {sin_imagen} carta(s) sin imagen ni UUID de Scryfall")

    return tam


def procesar(input_path: Path):
    es_csv    = input_path.suffix.lower() == ".csv"
    columns   = COLUMNS + (CSV_EXTRA_COLUMNS if (es_csv and INCLUDE_CSV_EXTRAS) else [])
    deck_name = input_path.stem

    # Si existen un .txt y un .csv con el mismo nombre, evitamos pisar el Excel
    stem = input_path.stem
    hermano = input_path.with_suffix(".txt" if es_csv else ".csv")
    if hermano.exists():
        stem = f"{stem}_{'csv' if es_csv else 'txt'}"
    xlsx_path = input_path.with_name(stem + ".xlsx")
    html_path = input_path.with_name(stem + ".html")

    print(f"\n{'='*54}")
    print(f"  Fuente: {input_path.name}  ({'CSV colección' if es_csv else 'TXT mazo'})")
    print(f"  Excel : {xlsx_path.name}")
    print(f"  HTML  : {html_path.name}")
    print(f"{'='*54}")

    cards = parse_input(input_path)
    if not cards:
        print("  No se encontraron cartas en el archivo. Se omite.")
        return

    # Deduplicación por clave precisa (id / set+número / nombre)
    uniques = {}
    for c in cards:
        uniques.setdefault(cache_key(c), c)
    total_copias = sum(c["qty"] for c in cards)
    print(f"  Entradas: {len(cards)}  |  Únicas: {len(uniques)}  |  Copias: {total_copias}\n")

    cache = build_cache(uniques)

    rows = [enrich(e, cache.get(cache_key(e)), columns) for e in cards]

    sin_datos = [r["name"] for r in rows if not r.get("oracle_text") and not r.get("type_line")]
    if sin_datos:
        print(f"\n  ⚠ {len(sin_datos)} carta(s) sin datos de Scryfall: "
              f"{', '.join(sorted(set(sin_datos))[:10])}"
              f"{'...' if len(set(sin_datos)) > 10 else ''}")

    groups = {o: [] for o in TYPE_ORDER}
    for row in rows:
        groups[sheet_for(row)].append(row)
    groups = {k: v for k, v in groups.items() if v}

    # ── Excel ──────────────────────────────────────────────────────────────────
    print("\n  Generando Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_summary(wb, groups)
    write_mana_curve(wb, rows)

    for sname, srows in groups.items():
        ws = wb.create_sheet(sname)
        ws.append([c.replace("_"," ").title() for c in columns])
        apply_header(ws, TYPE_COLORS.get(sname, TYPE_COLORS["Other"]))
        for row in srows:
            ws.append([row.get(c, "") for c in columns])
        style_data_rows(ws, columns)
        set_col_widths(ws, columns)
        ws.row_dimensions[1].height = 22

    wb.save(xlsx_path)
    print(f"  Excel guardado: {xlsx_path.name}")

    # ── HTML ───────────────────────────────────────────────────────────────────
    # Elegimos plantilla: un inventario y un mazo necesitan reportes distintos
    if MODO_FORZADO:
        modo = MODO_FORZADO
    elif es_csv and total_copias >= COLECCION_MIN_CARTAS:
        modo = "coleccion"
    else:
        modo = "mazo"

    if REPO_WEB:
        print("  Exportando datos para la web...")
        carpetas_img = [input_path.parent, Path(__file__).parent]
        exportar_web(
            rows, REPO_WEB, deck_name, sheet_for=sheet_for,
            titulo=TITULO_COLECCION or deck_name, subtitulo=SUBTITULO,
            banner="assets/banner.jpg" if (REPO_WEB / "assets" / "banner.jpg").exists() else "",
        )

    print(f"  Generando HTML (modo {modo})...")
    if modo == "coleccion":
        # Buscamos logo/banner: primero lo configurado, si no, autodetección
        carpetas = [input_path.parent, Path(__file__).parent]
        logo   = resolver_imagen(LOGO, carpetas)   if LOGO   else buscar_imagen_auto("logo", carpetas)
        banner = resolver_imagen(BANNER, carpetas) if BANNER else buscar_imagen_auto("banner", carpetas)
        generate_html_coleccion(rows, deck_name, html_path, sheet_for=sheet_for,
                                logo=logo, banner=banner,
                                titulo=TITULO_COLECCION, subtitulo=SUBTITULO,
                                banner_modo=BANNER_MODO)
    else:
        generate_html(rows, deck_name, html_path)
    print(f"  HTML  guardado: {html_path.name}")

    # ── Resumen ────────────────────────────────────────────────────────────────
    print(f"\n  Resumen:")
    for sname, srows in groups.items():
        count = sum(r["qty"] for r in srows if isinstance(r["qty"],int))
        print(f"    {sname:<15} {count:>3} cartas")
    print(f"    {'TOTAL':<15} {sum(r['qty'] for r in rows if isinstance(r['qty'],int)):>3} cartas")
    print()


# ── Main ──────────────────────────────────────────────────────────────────────

def resolver_argumentos(args: list[str]) -> list[Path]:
    """
    Convierte los argumentos de línea de comandos en rutas de archivo reales.
    Acepta rutas absolutas, relativas, carpetas y comodines (*.csv, Seph*.txt).
    """
    rutas = []
    for arg in args:
        p = Path(arg)

        # Carpeta → todos los .txt y .csv que contenga
        if p.is_dir():
            rutas.extend(sorted(list(p.glob("*.txt")) + list(p.glob("*.csv"))))
            continue

        # Archivo existente
        if p.exists():
            rutas.append(p)
            continue

        # Comodín sin expandir (CMD no expande * como bash)
        if any(ch in arg for ch in "*?"):
            carpeta = p.parent if str(p.parent) != "" else Path(".")
            rutas.extend(sorted(carpeta.glob(p.name)))
            continue

        print(f"  ⚠ No existe: {arg}")

    # Filtramos extensiones válidas y quitamos duplicados conservando el orden
    vistos, validas = set(), []
    for r in rutas:
        if r.suffix.lower() not in (".txt", ".csv"):
            print(f"  ⚠ Extensión no soportada (se omite): {r.name}")
            continue
        clave = str(r.resolve()).lower()
        if clave not in vistos:
            vistos.add(clave)
            validas.append(r)
    return validas


def main():
    global MODO_FORZADO
    base_dir = Path(__file__).parent
    edh_dir  = base_dir / "EDH"

    # Forzar plantilla HTML:  --coleccion  |  --mazo
    if "--coleccion" in sys.argv:
        MODO_FORZADO = "coleccion"
    elif "--mazo" in sys.argv:
        MODO_FORZADO = "mazo"

    # --web <ruta_al_repo>: exporta los JSON del sitio estático
    global REPO_WEB
    if "--web" in sys.argv:
        i = sys.argv.index("--web")
        if i + 1 < len(sys.argv) and not sys.argv[i + 1].startswith("--"):
            REPO_WEB = Path(sys.argv[i + 1])
            sys.argv.pop(i + 1)
        else:
            print("  ⚠ --web necesita la ruta del repositorio. Ejemplo:")
            print('     python manabox_to_excel.py Collection.csv --web "G:\\...\\ColeccionWeb"')
            sys.exit(1)

    # ── Modo 0: diagnóstico de conexión ───────────────────────────────────────
    if "--diagnostico" in sys.argv or "--diagnostic" in sys.argv:
        diagnostico()
        input("\nPresiona Enter para cerrar...")
        return

    # ── Modo 1: rutas pasadas por línea de comandos ───────────────────────────
    # python manabox_to_excel.py "C:\ruta\Sephiroth.txt"
    # python manabox_to_excel.py mazo1.txt mazo2.txt Collection.csv
    # python manabox_to_excel.py "C:\ruta\carpeta"
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if args:
        archivos = resolver_argumentos(args)
        if not archivos:
            print("\nNo se pudo procesar ningún archivo. Verifica las rutas.")
            sys.exit(1)
        for archivo in archivos:
            procesar(archivo)
        print(f"\n{len(archivos)} archivo(s) procesado(s).")
        return

    # ── Modo 2: carpeta EDH junto al script (menú interactivo / --todos) ──────
    if not edh_dir.exists():
        edh_dir.mkdir()
        print(f"Carpeta creada: {edh_dir}")
        print("Coloca tus archivos .txt o .csv de Manabox en la carpeta EDH y vuelve a correr.")
        sys.exit(0)

    archivos = sorted(
        list(edh_dir.glob("*.txt")) + list(edh_dir.glob("*.csv")),
        key=lambda p: (p.stem.lower(), p.suffix)
    )
    if not archivos:
        print(f"No se encontraron archivos .txt ni .csv en: {edh_dir}")
        print("Exporta tu mazo (.txt) o tu colección (.csv) desde Manabox,")
        print("guárdalo en EDH\\ y vuelve a correr.")
        sys.exit(0)

    if len(sys.argv) > 1 and sys.argv[1] == "--todos":
        for archivo in archivos:
            procesar(archivo)
        print("Todos los archivos procesados.")
        return

    print("\n╔══════════════════════════════════════════╗")
    print("║  Manabox → Excel + HTML  (Scryfall) v2.1 ║")
    print("╚══════════════════════════════════════════╝\n")
    print("  Archivos disponibles en EDH:\n")
    for i, archivo in enumerate(archivos, start=1):
        etiqueta = "CSV colección" if archivo.suffix.lower() == ".csv" else "TXT mazo"
        print(f"    [{i}] {archivo.stem:<30} ({etiqueta})")
    print(f"    [0] Procesar TODOS\n")

    try:
        opcion = int(input("  Selecciona una opción: ").strip())
    except ValueError:
        print("Opción inválida.")
        sys.exit(1)

    if opcion == 0:
        for archivo in archivos:
            procesar(archivo)
    elif 1 <= opcion <= len(archivos):
        procesar(archivos[opcion - 1])
    else:
        print("Opción fuera de rango.")
        sys.exit(1)

    input("\nPresiona Enter para cerrar...")


if __name__ == "__main__":
    main()
